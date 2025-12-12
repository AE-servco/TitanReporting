import datetime as dt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule

import modules.helpers as helpers


def formatted_cell(worksheet, row: int, col: int, val = None, font: Font | None=None, border: Border | None=None, number_format: str | None=None, fill: PatternFill | None=None):
    if val or val == 0:
        cell = worksheet.cell(row, col, val)
    else:
        cell = worksheet.cell(row, col)
    if font:
        cell.font = font
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill

    return cell

def build_workbook(
    jobs_by_tech: dict[str, list[dict]],
    end_date: dt.date
) -> bytes:
    
    CATEGORY_ORDER = {
        'wk_complete_paid': 'COMPLETED & PAID JOBS', 
        'wkend_complete_paid': 'WEEKEND COMPLETED & PAID JOBS', 
        'ah_complete_paid': 'AFTERHOURS COMPLETED & PAID JOBS', 
        'prev': 'PREVIOUS JOBS COMPLETED & PAID (COMMISSION) - Modoras Team Please do ADD to THIS SECTION or AMEND/TOUCH', 
        'wk_complete_unpaid': 'CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
        'wkend_complete_unpaid': 'WEEKEND CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
        'ah_complete_unpaid': 'AFTERHOURS CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
        'wk_wo': 'CURRENT WORK ORDERS (AWAITING PAYMENT)', 
        'wkend_wo': 'WEEKEND WORK ORDERS (AWAITING PAYMENT)', 
        'ah_wo': 'AFTERHOURS WORK ORDERS (AWAITING PAYMENT)', 
        'wk_unsuccessful': 'UNSUCCESSFUL JOBS',
        'wkend_unsuccessful': 'WEEKEND UNSUCCESSFUL JOBS',
        'ah_unsuccessful': 'AFTERHOURS UNSUCCESSFUL JOBS',
        'wk_uncategorised': 'WEEK UNCATEGORISED',
        'wkend_uncategorised': 'WEEKEND UNCATEGORISED',
        'ah_uncategorised': 'AFTERHOURS UNCATEGORISED',
    }

    # month = end_date.month
    # days_in_month = end_date.day

    dates = {
        'monday': end_date - dt.timedelta(days=6),
        'tuesday': end_date - dt.timedelta(days=5),
        'wednesday': end_date - dt.timedelta(days=4),
        'thursday': end_date - dt.timedelta(days=3),
        'friday': end_date - dt.timedelta(days=2),
        'saturday': end_date - dt.timedelta(days=1),
    }
    
    date_strs = {day: date.strftime("%d/%m/%Y") for day, date in dates.items()}
    # monday_str = monday.strftime("%d/%m/%Y")
    # tuesday_str = tuesday.strftime("%d/%m/%Y")
    # wednesday_str = wednesday.strftime("%d/%m/%Y")
    # thursday_str = thursday.strftime("%d/%m/%Y")
    # friday_str = friday.strftime("%d/%m/%Y")
    # saturday_str = saturday.strftime("%d/%m/%Y")

    cats_count_for_total = [
        'wk_complete_paid',
        'wkend_complete_paid',
        'wk_complete_unpaid',
        'wkend_complete_unpaid',
        'wk_wo',
        'wkend_wo',
        'ah_complete_paid',
        'ah_complete_unpaid',
        'ah_wo',
    ]

    cats_count_awaiting_pay_wk = [
        'wk_complete_unpaid',
        'wk_wo',
    ]

    cats_count_awaiting_pay_wkend = [
        'wkend_complete_unpaid',
        'wkend_wo',
        'ah_complete_unpaid',
        'ah_wo',
    ]

    cats_count_for_potential_wk = [
        'wk_complete_paid',
        'wk_complete_unpaid',
        'wk_wo',
        'wk_unsuccessful',
    ]

    cats_count_for_potential_wkend = [
        'wkend_complete_paid',
        'wkend_complete_unpaid',
        'wkend_wo',
        'wkend_unsuccessful',
    ]

    cats_count_for_unsuccessful = [
        'wk_unsuccessful',
        'wkend_unsuccessful',
        'ah_unsuccessful',
    ]

    # =========================== STYLING ===========================

    thin_border = Side(style='thin', color='000000') # Black thin border
    med_border = Side(style='medium', color='000000') # Black medium border
    thick_border = Side(style='thick', color='000000') # Black thick border
    double_border = Side(style='double', color='000000')

    yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
    blue_fill = PatternFill(patternType='solid', fgColor='3399F2')

    font_red = Font(color='FF0000')
    font_red_bold = Font(color='FF0000', bold=True)
    font_green_bold = Font(color='1F9C40', bold=True)
    font_green = Font(color='1F9C40')
    font_purple = Font(color='5F4F7D')
    font_bold = Font(bold=True)

    cell_border_full = Border(top=thin_border, bottom=thin_border, right=thin_border, left=thin_border)
    border_double_bottom = Border(bottom=double_border)

    accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    percentage_format = '0.00%'

    cell_border = {
        'topleft': Border(left=med_border, top=med_border),
        'topright': Border(right=med_border, top=med_border),
        'topleftright': Border(left=med_border, top=med_border, right=med_border),
        'bottomleft': Border(left=med_border, bottom=med_border),
        'bottomright': Border(right=med_border, bottom=med_border),
        'bottomtop': Border(top=med_border, bottom=med_border),
        'bottomtop_double': Border(top=med_border, bottom=double_border),
        'bottomtopright': Border(top=med_border, bottom=med_border, right=med_border),
        'bottomleftright': Border(left=med_border, bottom=med_border, right=med_border),
        'bottom': Border(bottom=med_border),
        'top': Border(top=med_border),
        'right': Border(right=med_border),
        'left': Border(left=med_border),
        'leftright': Border(left=med_border, right=med_border),
    }

    align_center = Alignment(horizontal='center')

    # =========================== WORKBOOK ===========================

    wb = Workbook()
    
    first_sheet_created = False

    col_offset = 1

    for tech in sorted(jobs_by_tech.keys()):

        job_cats = jobs_by_tech[tech]
        if not first_sheet_created:
            ws = wb.active
            ws.title = tech[:-1]
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title=tech[:-1])

        sales_color = '3da813'
        installer_color = 'e3b019'
        unknown_color = '8093f2'

        if tech[-1] == 'S':
            ws.sheet_properties.tabColor = sales_color
        elif tech[-1] == 'I':
            ws.sheet_properties.tabColor = installer_color
        else:
            ws.sheet_properties.tabColor = unknown_color

        summary_top_row = 1

        # ================ SUMMARY ==================
        formatted_cell(ws, summary_top_row, col_offset + 1, 'WEEKLY COMMISSION', font = font_bold)
        formatted_cell(ws, summary_top_row, col_offset + 5, 'WEEK ENDING', font = font_red_bold)
        formatted_cell(ws, summary_top_row, col_offset + 6, end_date.strftime("%d/%m/%Y"), font = font_red_bold)

        # box 1
        formatted_cell(ws, summary_top_row + 2, col_offset + 1, 'BOOKED JOBS',font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 2, '=C4+C5', font = font_bold, border = cell_border['topright'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 1, 'SUCCESSFUL', font = font_bold, border = cell_border['left'])
        # formatted_cell(ws, summary_top_row + 11, col_offset + 17, '=SUM(K11:O11) + SUM(K17:O17) + SUM(K23:O23) + SUM(K29:O29) + SUM(K35:O35) + SUM(K41:O41)', font = font_bold, border = cell_border['bottomleft'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 3, col_offset + 2, '=SUM(K12:O12) + SUM(K18:O18) + SUM(K24:O24) + SUM(K30:O30) + SUM(K36:O36) + SUM(K42:O42)', font = font_bold, border = cell_border['right'])
        formatted_cell(ws, summary_top_row + 4, col_offset + 1, 'UNSUCCESSFUL', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 4, col_offset + 2, '=SUM(K13:O13) + SUM(K19:O19) + SUM(K25:O25) + SUM(K31:O31) + SUM(K37:O37) + SUM(K43:O43)', font = font_bold, border = cell_border['right'])
        formatted_cell(ws, summary_top_row + 5, col_offset + 1, 'SUCCESSFUL (%)', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 5, col_offset + 2, '=C4/(C4+C5)', font = font_bold, border = cell_border['right'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 6, col_offset + 1, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 6, col_offset + 2, border = cell_border['right'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 1, 'AVERAGE SALE', font = font_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 2, '=R12/C4', font = font_bold, border = cell_border['bottomright'], number_format=accounting_format)
        
        # box 2
        formatted_cell(ws, summary_top_row + 2, col_offset + 4, 'WEEKLY TARGET', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 5, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 6, border = cell_border['topright'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 4, 'Tier 1', border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 5, '<$25000')
        formatted_cell(ws, summary_top_row + 3, col_offset + 6, '=0.05', border = cell_border['right'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 4, 'Tier 2', border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 4, col_offset + 5, '>=$25000', border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 4, col_offset + 6, '=0.1', border = cell_border['bottomright'], number_format=percentage_format)

        # box 3
        formatted_cell(ws, summary_top_row + 6, col_offset + 5, 'ACTUAL', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 6, col_offset + 6, 'POTENTIAL', font = font_bold, border = cell_border['topright'])
        formatted_cell(ws, summary_top_row + 6, col_offset + 7, 'Exc. SUPER', font = font_green_bold)
        formatted_cell(ws, summary_top_row + 7, col_offset + 4, 'NET PROFIT', font = font_bold, border = cell_border['topleft'])
        # Some down below

        formatted_cell(ws, summary_top_row + 8, col_offset + 4, 'UNLOCKED', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 5, '=IF(R12>=25000,G5,G4)', border = cell_border['left'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 9, col_offset + 4, 'COMMISSION - PAY OUT', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 9, col_offset + 5, '=F8*F9', border = cell_border['left'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 9, col_offset + 7, '=F10/1.12', font = font_green_bold, number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 10, col_offset + 4, 'EMERGENCY', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 10, col_offset + 5, '=S12', border = cell_border['left'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 11, col_offset + 4, 'EMERGENCY - PAY OUT', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 11, col_offset + 5, '=F11*0.25', border = cell_border['left'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 11, col_offset + 7, '=F12/1.12', font = font_green_bold, number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 12, col_offset + 4, 'PREV. WEEK', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 12, col_offset + 5, 0, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 13, col_offset + 4, 'PREV. WEEK - PAY OUT', font = font_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 13, col_offset + 5, '=F13*0.05', border = cell_border['bottomleft'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 13, col_offset + 7, '=F14/1.12', font = font_green_bold, number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 14, col_offset + 4, '5 Star Review', font = font_green_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 14, col_offset + 5, '=F16*50', border = cell_border['left'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 15, col_offset + 4, '5 Star Notes', font = font_green_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 15, col_offset + 5, '=T4', border = cell_border['left'])

        formatted_cell(ws, summary_top_row + 8, col_offset + 6, '==IF(G8>=25000,G5,G4)', font=font_red, border = cell_border['right'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 9, col_offset + 6, '=G8*G9', font = font_red, border = cell_border['right'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 11, col_offset + 6, border = cell_border['right'])
        formatted_cell(ws, summary_top_row + 12, col_offset + 6, 0, font = font_red, border = cell_border['right'])
        formatted_cell(ws, summary_top_row + 13, col_offset + 6, '==G13*0.05', font = font_red, border = cell_border['bottomright'], number_format=accounting_format)
        
        #box 4
        formatted_cell(ws, summary_top_row + 1, col_offset + 10, 'PHOTOS', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 11, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 12, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 13, 'QUOTE', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 14, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 15, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 16, 'INVOICE', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 17, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 18, border = cell_border['topright'])
        formatted_cell(ws, summary_top_row + 1, col_offset + 19, border = cell_border['topright'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 10, 'BEFORE', font = font_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 11, 'AFTER', font = font_bold, border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 12, 'RECEIPT', font = font_bold, border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 13, 'DESCRIPTION', font = font_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 14, 'SIGNED', font = font_bold, border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 15, 'EMAILED', font = font_bold, border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 16, 'DESCRIPTION', font = font_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 17, 'SIGNED', font = font_bold, border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 18, 'EMAILED', font = font_bold, border = cell_border['bottom'])
        formatted_cell(ws, summary_top_row + 2, col_offset + 19, '5 Star Review', font = font_bold, border = cell_border['bottomleftright'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 9, 'TAKEN', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 4, col_offset + 9, '%', font = font_bold, border = cell_border['bottomleft'])
        formatted_cell(ws, summary_top_row + 4, col_offset + 19, border = cell_border['bottomleftright'])
        # inputs below because they need to know row numbers

        # box 5
        formatted_cell(ws, summary_top_row + 7, col_offset + 10, 'DAILY NET PROFIT', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 11, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 12, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 13, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 14, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 15, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 16, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 17, 'WEEKDAY', font = font_red_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 7, col_offset + 18, 'WEEKEND', font = font_red_bold, border = cell_border['topleftright'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 10, 'MON', font = font_bold, border = cell_border['bottomtop_double'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 11, 'TUE', font = font_bold, border = cell_border['bottomtop_double'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 12, 'WED', font = font_bold, border = cell_border['bottomtop_double'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 13, 'THU', font = font_bold, border = cell_border['bottomtop_double'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 14, 'FRI', font = font_bold, border = cell_border['bottomtop_double'])
        # formatted_cell(ws, summary_top_row + 8, col_offset + 15, 'Total', font = font_bold, border = cell_border['top'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 15, 'SAT', font = font_green_bold, border = cell_border['bottomtop_double'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 16, 'SUN', font = font_green_bold, border = cell_border['bottomtop_double'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 17, 'Awaiting Payment', font = font_red_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 8, col_offset + 18, 'Awaiting Payment', font = font_red_bold, border = cell_border['topleftright'])
        ##
        formatted_cell(ws, summary_top_row + 10, col_offset + 17, 'Total Payable', font = font_bold, border = cell_border['topleft'])
        formatted_cell(ws, summary_top_row + 10, col_offset + 18, 'Total Payable', font = font_bold, border = cell_border['topleftright'])

        formatted_cell(ws, summary_top_row + 11, col_offset + 17, '=SUM(K11:O11) + SUM(K17:O17) + SUM(K23:O23) + SUM(K29:O29) + SUM(K35:O35) + SUM(K41:O41)', font = font_bold, border = cell_border['bottomleft'], number_format=accounting_format)
        formatted_cell(ws, summary_top_row + 11, col_offset + 18, '=SUM(P11:Q11) + SUM(P17:Q17) + SUM(P23:Q23) + SUM(P29:Q29) + SUM(P35:Q35) + SUM(P41:Q41)', font = font_bold, border = cell_border['bottomright'], number_format=accounting_format)

        # formatted_cell(ws, summary_top_row + 10, col_offset + 17, '=SUM(K11:O11)', font = font_bold, border = cell_border['top'], number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 16, col_offset + 17, '=SUM(K17:O17)', font = font_bold, border = cell_border['top'], number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 22, col_offset + 17, '=SUM(K23:O23)', font = font_bold, border = cell_border['top'], number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 28, col_offset + 17, '=SUM(K29:O29)', font = font_bold, border = cell_border['top'], number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 34, col_offset + 17, '=SUM(K35:O35)', font = font_bold, border = cell_border['top'], number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 40, col_offset + 17, '=SUM(K41:O41)', font = font_bold, border = cell_border['top'], number_format=accounting_format)
        ##


        formatted_cell(ws, summary_top_row + 11, col_offset + 9, 'Successful')
        formatted_cell(ws, summary_top_row + 12, col_offset + 9, 'Unsuccessful')
        formatted_cell(ws, summary_top_row + 13, col_offset + 9, 'Success rate')
        formatted_cell(ws, summary_top_row + 14, col_offset + 9, 'Avg sale')

        # # box 6 - management summary
        # formatted_cell(ws, summary_top_row + 15, col_offset + 10, 'Management Summary - SALES ONLY - COLUMN F', font = font_bold, border = cell_border['topleft'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 11, border = cell_border['top'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 12, border = cell_border['top'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 13, border = cell_border['top'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 14, border = cell_border['top'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 15, border = cell_border['top'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 16, border = cell_border['top'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 17, 'WEEKDAY', font = font_red_bold, border = cell_border['topleft'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 15, col_offset + 18, 'WEEKEND', font = font_red_bold, border = cell_border['topleftright'], fill=blue_fill)
        # formatted_cell(ws, summary_top_row + 16, col_offset + 10, 'MON', font = font_bold, border = cell_border['topleft'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 11, 'TUE', font = font_bold, border = cell_border['top'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 12, 'WED', font = font_bold, border = cell_border['top'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 13, 'THU', font = font_bold, border = cell_border['top'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 14, 'FRI', font = font_bold, border = cell_border['top'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 15, 'Total', font = font_bold, border = cell_border['top'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 16, 'SAT', font = font_green_bold, border = cell_border['topleft'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 17, 'Awaiting Payment', font = font_red_bold, border = cell_border['topleft'])
        # formatted_cell(ws, summary_top_row + 16, col_offset + 18, 'Awaiting Payment', font = font_red_bold, border = cell_border['topleftright'])
        # ##
        # formatted_cell(ws, summary_top_row + 17, col_offset + 15, '=SUM(K18:O18)', font = font_bold, border = cell_border['bottom'], number_format=accounting_format)#
        # ##

        # formatted_cell(ws, summary_top_row + 18, col_offset + 9, 'Successful')
        # formatted_cell(ws, summary_top_row + 19, col_offset + 9, 'Unsuccessful')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 9, 'Success rate')
        # formatted_cell(ws, summary_top_row + 21, col_offset + 9, 'Avg sale')

        # formatted_cell(ws, summary_top_row + 18, col_offset + 10, '=K11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 10, '=K12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 10, '=K13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 10, '=K18/K19', number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 18, col_offset + 11, '=L11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 11, '=L12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 11, '=L13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 11, '=L18/L19', number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 18, col_offset + 12, '=M11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 12, '=M12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 12, '=M13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 12, '=M18/M19', number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 18, col_offset + 13, '=N11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 13, '=N12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 13, '=N13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 13, '=N18/N19', number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 18, col_offset + 14, '=O11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 14, '=O12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 14, '=O13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 14, '=O18/O19', number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 18, col_offset + 15, '=P11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 15, '=P12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 15, '=P13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 15, '=P18/P19', number_format=accounting_format)
        # formatted_cell(ws, summary_top_row + 18, col_offset + 16, '=Q11') 
        # formatted_cell(ws, summary_top_row + 19, col_offset + 16, '=Q12')
        # formatted_cell(ws, summary_top_row + 20, col_offset + 16, '=Q13', number_format=percentage_format)
        # formatted_cell(ws, summary_top_row + 21, col_offset + 16, '=Q18/Q19', number_format=accounting_format)

        # curr_row += 1

        # ws.cell(curr_row,10, 'PHOTOS').font = font_bold
        # ws.cell(curr_row,10).border = cell_border['topleft']
        # ws.cell(curr_row,11).border = cell_border['top']
        # ws.cell(curr_row,12).border = cell_border['top']
        # ws.cell(curr_row,13, 'QUOTE').font = font_bold
        # ws.cell(curr_row,13).border = cell_border['topleft']
        # ws.cell(curr_row,14).border = cell_border['top']
        # ws.cell(curr_row,15).border = cell_border['top']
        # ws.cell(curr_row,16, 'INVOICE').font = font_bold
        # ws.cell(curr_row,16).border = cell_border['topleft']
        # ws.cell(curr_row,17).border = cell_border['top']
        # ws.cell(curr_row,18).border = cell_border['topright']
        # ws.cell(curr_row,19).border = cell_border['topright']

        # curr_row += 1

        # ws.cell(curr_row,4, 'WEEKLY TARGET').font = font_bold
        # ws.cell(curr_row,4).border = cell_border['topleft']
        # ws.cell(curr_row,5).border = cell_border['top']
        # ws.cell(curr_row,6).border = cell_border['topright']
        # ws.cell(curr_row,10, 'BEFORE').font = font_bold

        curr_row = 50
        # ================ HEADERS ==================
        ws.cell(curr_row, col_offset + 1, 'JOB DETAILS').border = cell_border['topleft']
        ws.cell(curr_row, col_offset + 2).border = cell_border['top']
        ws.cell(curr_row, col_offset + 3).border = cell_border['top']
        ws.cell(curr_row, col_offset + 4).border = cell_border['top']
        ws.cell(curr_row, col_offset + 5, 'JOB AMOUNT').border = cell_border['topleft']
        ws.cell(curr_row, col_offset + 6).border = cell_border['top']
        ws.cell(curr_row, col_offset + 7).border = cell_border['top']
        ws.cell(curr_row, col_offset + 8).border = cell_border['top']
        ws.cell(curr_row, col_offset + 9).border = cell_border['top']
        ws.cell(curr_row, col_offset + 10).border = cell_border['topleft']
        ws.cell(curr_row, col_offset + 11, 'PHOTOS').border = cell_border['topleft']
        ws.cell(curr_row, col_offset + 12).border = cell_border['top']
        ws.cell(curr_row, col_offset + 13).border = cell_border['top']
        ws.cell(curr_row, col_offset + 14, 'QUOTE').border = cell_border['topleft']
        ws.cell(curr_row, col_offset + 15).border = cell_border['top']
        ws.cell(curr_row, col_offset + 16).border = cell_border['top']
        ws.cell(curr_row, col_offset + 17, 'INVOICE').border = cell_border['topleft']
        ws.cell(curr_row, col_offset + 18).border = cell_border['top']
        ws.cell(curr_row, col_offset + 19).border = cell_border['top']
        ws.cell(curr_row, col_offset + 20).border = cell_border['top']
        ws.cell(curr_row, col_offset + 21, 'NOTES').border = cell_border['topright']

        curr_row += 1

        ws.cell(curr_row, col_offset + 1, 'JOB STATUS').border = cell_border['bottomleft']
        ws.cell(curr_row, col_offset + 2, 'DATE').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 3, 'JOB #').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 4, 'SUBURB').border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 5, 'AMOUNT EXC. GST').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 6, 'MATERIALS').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 7, 'MERCHANT FEES').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 8, 'NET PROFIT').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 9, 'PAID').border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 10, 'DOC CHECK DONE').border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 11, 'BEFORE').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 12, 'AFTER').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 13, 'RECEIPT').border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 14, 'DESCRIPTION').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 15, 'SIGNED').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 16, 'EMAILED').border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 17, 'DESCRIPTION').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 18, 'SIGNED').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 19, 'EMAILED').border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 20, '5 Star Review').border = cell_border['bottom']
        ws.cell(curr_row, col_offset + 21).border = cell_border['bottomright']
        ws.cell(curr_row, col_offset + 22, 'EFTPOS').border = cell_border['bottomtop']
        ws.cell(curr_row, col_offset + 23, 'CASH').border = cell_border['bottomtop']
        ws.cell(curr_row, col_offset + 24, 'Payment Plan').border = cell_border['bottomtopright']
        
        curr_row += 1
        # =========================================

        cat_row_info = {}

        for cat, cat_text in CATEGORY_ORDER.items():
            cat_font = None
            if cat in ['wk_complete_unpaid', 'wk_wo', 'wkend_complete_unpaid', 'wkend_wo']:
                cat_font = font_green
            formatted_cell(ws, curr_row, col_offset + 1,cat_text, font=cat_font)
            curr_row += 1
            cat_row_start = curr_row

            if cat == 'prev':
                for row in ws[f'A{curr_row-1}:X{curr_row+7}']:
                    for cell in row:
                        cell.fill = yellow_fill
                for i in range(1,9):
                    ws.cell(curr_row, col_offset + 1, i)
                    curr_row += 1

            else:
                jobs = job_cats.get(cat, [])
                if not jobs:
                    curr_row += 1
                job_count = 1
                curr_date = ""
                for job in jobs:
                    if job['first_appt_start_str'] != curr_date:
                        job_border = cell_border['top']
                    else:
                        job_border = None
                    curr_date = job['first_appt_start_str']

                    if job['complaint_tag_present']:
                        formatted_cell(ws, curr_row, col_offset, 'COMPLAINT', font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 1, job_count, font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 2, job['first_appt_start_str'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 3, int(job['num']), font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 4, job['suburb'], font=cat_font, border = job_border)
                    if not job['unsuccessful']:
                        formatted_cell(ws, curr_row, col_offset + 5, job['inv_subtotal'], font=cat_font, border = job_border, number_format=accounting_format)
                        formatted_cell(ws, curr_row, col_offset + 6, job['inv_subtotal']*0.2, font=cat_font, border = job_border, number_format=accounting_format)
                        formatted_cell(ws, curr_row, col_offset + 6, font=cat_font, border = job_border).comment = Comment(job['summary'], "automation")
                    else:
                        formatted_cell(ws, curr_row, col_offset + 5, job['open_est_subtotal'], font=cat_font, border = job_border, number_format=accounting_format)
                        formatted_cell(ws, curr_row, col_offset + 6, "", font=cat_font, border = job_border)
                    # 7
                    formatted_cell(ws, curr_row, col_offset + 7, "", font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 8, f"={get_column_letter(col_offset + 5)}{curr_row}-{get_column_letter(col_offset + 6)}{curr_row}-{get_column_letter(col_offset + 7)}{curr_row}", font=cat_font, border = job_border, number_format=accounting_format)
                    formatted_cell(ws, curr_row, col_offset + 9, job['payment_types'], font=cat_font, border = job_border)
                    # 10 TODO: all doc checks complete
                    doc_check_complete_col = f'=IF(OR({", ".join([f"{get_column_letter(col_offset + i)}{curr_row}=0" for i in range(11,20)])}), "N","Y")'
                    formatted_cell(ws, curr_row, col_offset + 10, doc_check_complete_col, font=cat_font, border = job_border).alignment = align_center
                    formatted_cell(ws, curr_row, col_offset + 11, job['Before Photo'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 12, job['After Photo'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 13, job['Receipt Photo'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 14, job['Quote Description'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 15, job['Quote Signed'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 16, job['Quote Emailed'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 17, job['Invoice Description'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 18, job['Invoice Signed'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 19, job['Invoice Emailed'], font=cat_font, border = job_border)
                    formatted_cell(ws, curr_row, col_offset + 20, job['5 Star Review'], font=cat_font, border = job_border)

                    formatted_cell(ws, curr_row, col_offset + 25, f"=ROUND(F{curr_row}*1.1,2)", font=font_purple)
                    formatted_cell(ws, curr_row, col_offset + 26, f"=ROUND(Z{curr_row} - W{curr_row} - X{curr_row} - Y{curr_row},2)", font=font_purple)
                    
                    payments = job.get('payment_amt')

                    if type(payments) == str:
                        payments = payments.split(', ')
                        
                        # All payment types:
                        # {
                            # 'AMEX',
                            # 'Applied Payment for AR',
                            # 'Cash',
                            # 'Check',
                            # 'Credit Card',
                            # 'EFT/Bank Transfer',
                            # 'Humm - Finance Fee',
                            # 'Humm Payment Plan',
                            # 'Imported Default Credit Card',
                            # 'MasterCard',
                            # 'Payment Plan',
                            # 'Payment Plan - Fee',
                            # 'Processed in ServiceM8',
                            # 'Refund (check)',
                            # 'Refund (credit card)',
                            # 'Visa'
                        # }
                        
                        p_types = {
                            'Cr': [],
                            'Ca': [],
                            'EF': [],
                        }
                        for p in payments:
                            try:
                                p_types[p[:2]].append(p[2:])
                            except:
                                continue
                        if p_types['EF'] or p_types['Cr']:
                            formatted_cell(ws, curr_row, col_offset + 22, f"={'+'.join(p_types['EF'] + p_types['Cr'])}", font=cat_font, number_format=accounting_format)
                        if p_types['Ca']:
                            formatted_cell(ws, curr_row, col_offset + 23, f"={'+'.join(p_types['Ca'])}", font=cat_font, number_format=accounting_format)
                    curr_row += 1
                    job_count += 1
            

            # add 5 rows for any extras
            curr_row += 5

            # Fill in info 
            cat_row_info[cat] = (cat_row_start, curr_row-1)

            # Doc check formatting 0s to red
            dc_start_col_letter = get_column_letter(col_offset + 10)
            dc_end_col_letter = get_column_letter(col_offset + 20)
            # dc_range = "A1:A100"
            dc_range = f'{dc_start_col_letter}{cat_row_start}:{dc_end_col_letter}{curr_row-1}'
            ws.conditional_formatting.add(
                dc_range, 
                CellIsRule(operator="equal", formula=["0"], font=font_red)
            )

            # totals row
            amt_col = 5
            amt_letter = get_column_letter(col_offset + amt_col)
            materials_col = 6
            materials_letter = get_column_letter(col_offset + materials_col)
            merchantf_col = 7
            merchantf_letter = get_column_letter(col_offset + merchantf_col)
            profit_col = 8
            profit_letter = get_column_letter(col_offset + profit_col)
            ws.cell(curr_row, col_offset + amt_col, f"=SUM({amt_letter}{cat_row_start}:{amt_letter}{curr_row-1})").number_format = accounting_format
            ws.cell(curr_row, col_offset + materials_col, f"=SUM({materials_letter}{cat_row_start}:{materials_letter}{curr_row-1})").number_format = accounting_format
            ws.cell(curr_row, col_offset + merchantf_col, f"=SUM({merchantf_letter}{cat_row_start}:{merchantf_letter}{curr_row-1})").number_format = accounting_format
            ws.cell(curr_row, col_offset + profit_col, f"=SUM({profit_letter}{cat_row_start}:{profit_letter}{curr_row-1})").number_format = accounting_format
            ws.cell(curr_row, col_offset + amt_col).border = border_double_bottom
            ws.cell(curr_row, col_offset + materials_col).border = border_double_bottom
            ws.cell(curr_row, col_offset + merchantf_col).border = border_double_bottom
            ws.cell(curr_row, col_offset + profit_col).border = border_double_bottom

            if cat == 'wk_complete_paid':
                formatted_cell(ws, curr_row, col_offset + 11, f"=COUNT(L{cat_row_start}:L{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 12, f"=COUNT(M{cat_row_start}:M{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 13, f"=COUNT(N{cat_row_start}:N{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 14, f"=COUNT(O{cat_row_start}:O{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 15, f"=COUNT(P{cat_row_start}:P{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 16, f"=COUNT(Q{cat_row_start}:Q{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 17, f"=COUNT(R{cat_row_start}:R{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 18, f"=COUNT(S{cat_row_start}:S{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 19, f"=COUNT(T{cat_row_start}:T{curr_row-1})")
                formatted_cell(ws, curr_row, col_offset + 20, f"=COUNT(U{cat_row_start}:U{curr_row-1})")

                formatted_cell(ws, curr_row+1, col_offset + 11, f"=SUM(L{cat_row_start}:L{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 12, f"=SUM(M{cat_row_start}:M{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 13, f"=SUM(N{cat_row_start}:N{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 14, f"=SUM(O{cat_row_start}:O{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 15, f"=SUM(P{cat_row_start}:P{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 16, f"=SUM(Q{cat_row_start}:Q{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 17, f"=SUM(R{cat_row_start}:R{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 18, f"=SUM(S{cat_row_start}:S{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 19, f"=SUM(T{cat_row_start}:T{curr_row-1})")
                formatted_cell(ws, curr_row+1, col_offset + 20, f"=SUM(U{cat_row_start}:U{curr_row-1})")
                curr_row += 1
            curr_row += 2
        
        # =SUMIF(B{cat_start}:B{cat_end}, date_str, E{cat_start}:E{cat_end}) -- sales
        # =SUMIF(B{cat_start}:B{cat_end}, date_str, H{cat_start}:H{cat_end}) -- profit
# monday_str

        # days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
        # profit_formulas = {day: '=' + ' + '.join([f'SUMIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{date_strs[day]}", I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_for_total]) for day in days}


        subtract_red_formula_wk = f'SUMIF(K{cat_row_info["wk_complete_paid"][0]}:K{cat_row_info["wk_complete_paid"][1]}, "N", I{cat_row_info["wk_complete_paid"][0]}:I{cat_row_info["wk_complete_paid"][1]})'
        formatted_cell(ws, summary_top_row + 7, col_offset + 5, f'=R12-R10-{subtract_red_formula_wk}', border = cell_border['topleft'], number_format=accounting_format)
        # subtract_red_formula_wkend = f'SUMIF(K{cat_row_info["wkend_complete_paid"][0]}:K{cat_row_info["wkend_complete_paid"][1]}, "N", I{cat_row_info["wkend_complete_paid"][0]}:I{cat_row_info["wkend_complete_paid"][1]})'

        dates_in_month = helpers.get_dates_in_month_datetime(end_date.year, end_date.month)
        profit_formulas = {day: '=' + ' + '.join([f'SUMIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{day.strftime("%d/%m/%Y")}", I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_for_total]) for day in dates_in_month}
        count_success_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{day.strftime("%d/%m/%Y")}")'for cat in cats_count_for_total]) for day in dates_in_month}
        count_unsuccess_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{day.strftime("%d/%m/%Y")}")'for cat in cats_count_for_unsuccessful]) for day in dates_in_month}

        start_row = 9
        reset_col = 10
        start_col = reset_col

        SUMMARY_COL_LENGTH = 0
        for day in dates_in_month:
            day_row = start_row
            day_of_week = day.weekday()
            if day_of_week == 0:
                # If monday, write the LHS words 
                formatted_cell(ws, summary_top_row + day_row + 2, col_offset + start_col - 1, 'Successful')
                formatted_cell(ws, summary_top_row + day_row + 2 + 1, col_offset + start_col - 1, 'Unsuccessful')
                formatted_cell(ws, summary_top_row + day_row + 2 + 2, col_offset + start_col - 1, 'Success rate')
                formatted_cell(ws, summary_top_row + day_row + 2 + 3, col_offset + start_col - 1, 'Avg sale')
            
            summary_font = font_bold
            if day_of_week in [5,6]:
                summary_font = font_green_bold

            curr_col = col_offset + start_col + day_of_week
            formatted_cell(ws, summary_top_row + day_row, curr_col, day.strftime("%d/%m"), font = summary_font, border = cell_border['top']) 
            day_row += 1
            formatted_cell(ws, summary_top_row + day_row, curr_col, profit_formulas[day], font = summary_font, number_format=accounting_format) 
            day_row += 1
            formatted_cell(ws, summary_top_row + day_row, curr_col, count_success_formulas[day]) 
            day_row += 1
            formatted_cell(ws, summary_top_row + day_row, curr_col, count_unsuccess_formulas[day])
            day_row += 1
            curr_col_letter = get_column_letter(curr_col)
            formatted_cell(ws, summary_top_row + day_row, curr_col, f'={curr_col_letter}{summary_top_row + day_row-2}/({curr_col_letter}{summary_top_row + day_row-2}+{curr_col_letter}{summary_top_row + day_row-1})', number_format=percentage_format)
            day_row += 1
            formatted_cell(ws, summary_top_row + day_row, curr_col, f'={curr_col_letter}{summary_top_row + day_row-4}/{curr_col_letter}{summary_top_row + day_row-3}', number_format=accounting_format)
            day_row += 1

            SUMMARY_COL_LENGTH = day_row - start_row
            if day_of_week == 6:
                # If sunday, push row to next block
                start_row += SUMMARY_COL_LENGTH

        # formatted_cell(ws, summary_top_row + 9, col_offset + 10, profit_formulas['monday'], font = font_bold, border = cell_border['bottomleft'], number_format=accounting_format) # These all rely on daily totals
        # formatted_cell(ws, summary_top_row + 9, col_offset + 11, profit_formulas['tuesday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format) # These all rely on daily totals
        # formatted_cell(ws, summary_top_row + 9, col_offset + 12, profit_formulas['wednesday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format) # These all rely on daily totals
        # formatted_cell(ws, summary_top_row + 9, col_offset + 13, profit_formulas['thursday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format) # These all rely on daily totals
        # formatted_cell(ws, summary_top_row + 9, col_offset + 14, profit_formulas['friday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format) # These all rely on daily totals
        # formatted_cell(ws, summary_top_row + 9, col_offset + 16, profit_formulas['saturday'], font = font_green_bold, border = cell_border['bottomleft'], number_format=accounting_format) # These all rely on daily totals

        # sales_formulas = {day: '=' + ' + '.join([f'SUMIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{date_strs[day]}", F{cat_row_info[cat][0]}:F{cat_row_info[cat][1]})'for cat in cats_count_for_total]) for day in days}

        # formatted_cell(ws, summary_top_row + 17, col_offset + 10, sales_formulas['monday'], font = font_bold, border = cell_border['bottomleft'], number_format=accounting_format) # These all rely on daily totals
        # formatted_cell(ws, summary_top_row + 17, col_offset + 11, sales_formulas['tuesday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format)#
        # formatted_cell(ws, summary_top_row + 17, col_offset + 12, sales_formulas['wednesday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format)#
        # formatted_cell(ws, summary_top_row + 17, col_offset + 13, sales_formulas['thursday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format)#
        # formatted_cell(ws, summary_top_row + 17, col_offset + 14, sales_formulas['friday'], font = font_bold, border = cell_border['bottom'], number_format=accounting_format)#
        # formatted_cell(ws, summary_top_row + 17, col_offset + 16, sales_formulas['saturday'], font = font_green_bold, border = cell_border['bottomleft'], number_format=accounting_format) # These all rely on daily totals

        # count_success_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{date_strs[day]}")'for cat in cats_count_for_total]) for day in days}

        # formatted_cell(ws, summary_top_row + 10, col_offset + 10, count_success_formulas['monday']) # Actual count of monday jobs
        # formatted_cell(ws, summary_top_row + 10, col_offset + 11, count_success_formulas['tuesday']) # Actual count of tuesday jobs
        # formatted_cell(ws, summary_top_row + 10, col_offset + 12, count_success_formulas['wednesday']) # Actual count of wednesday jobs
        # formatted_cell(ws, summary_top_row + 10, col_offset + 13, count_success_formulas['thursday']) # Actual count of thursday jobs
        # formatted_cell(ws, summary_top_row + 10, col_offset + 14, count_success_formulas['friday']) # Actual count of friday jobs
        # formatted_cell(ws, summary_top_row + 10, col_offset + 16, count_success_formulas['saturday']) # Actual count of saturday jobs

        # count_unsuccess_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{cat_row_info[cat][0]}:C{cat_row_info[cat][1]}, "{date_strs[day]}")'for cat in cats_count_for_unsuccessful]) for day in days}

        # # Unsuccessful counts
        # formatted_cell(ws, summary_top_row + 11, col_offset + 10, count_unsuccess_formulas['monday'])
        # formatted_cell(ws, summary_top_row + 11, col_offset + 11, count_unsuccess_formulas['tuesday'])
        # formatted_cell(ws, summary_top_row + 11, col_offset + 12, count_unsuccess_formulas['wednesday'])
        # formatted_cell(ws, summary_top_row + 11, col_offset + 13, count_unsuccess_formulas['thursday'])
        # formatted_cell(ws, summary_top_row + 11, col_offset + 14, count_unsuccess_formulas['friday'])
        # formatted_cell(ws, summary_top_row + 11, col_offset + 16, count_unsuccess_formulas['saturday'])

        end_of_comp_paid = cat_row_info['wk_complete_paid'][1]
        formatted_cell(ws, summary_top_row + 3, col_offset + 10, f'=L{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 11, f'=M{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 12, f'=N{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 13, f'=O{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 14, f'=P{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 15, f'=Q{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 16, f'=R{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 17, f'=S{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        formatted_cell(ws, summary_top_row + 3, col_offset + 18, f'=T{end_of_comp_paid+2}', font = font_bold, border = cell_border['left'])
        
        
        review_count_formula = '=' + ' + '.join([f'SUM(U{cat_row_info[cat][0]}:U{cat_row_info[cat][1]})'for cat in cats_count_for_total])
        formatted_cell(ws, summary_top_row + 3, col_offset + 19, review_count_formula, font = font_bold, border = cell_border['leftright']) # total 5 star reviews

        formatted_cell(ws, summary_top_row + 4, col_offset + 10, f'=L{end_of_comp_paid+2}/L{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 11, f'=M{end_of_comp_paid+2}/M{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 12, f'=N{end_of_comp_paid+2}/N{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 13, f'=O{end_of_comp_paid+2}/O{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 14, f'=P{end_of_comp_paid+2}/P{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 15, f'=Q{end_of_comp_paid+2}/Q{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 16, f'=R{end_of_comp_paid+2}/R{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 17, f'=S{end_of_comp_paid+2}/S{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)
        formatted_cell(ws, summary_top_row + 4, col_offset + 18, f'=T{end_of_comp_paid+2}/T{end_of_comp_paid+1}', font = font_bold, border = cell_border['bottomleft'], number_format=percentage_format)

        

        wk_profit_awaiting_formula = '=' + ' + '.join([f'SUM(I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_awaiting_pay_wk])
        formatted_cell(ws, summary_top_row + 9, col_offset + 17, wk_profit_awaiting_formula, font = font_red_bold, border = cell_border['bottomleft'], number_format=accounting_format)
        wkend_profit_awaiting_formula = '=' + ' + '.join([f'SUM(I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_awaiting_pay_wkend])
        formatted_cell(ws, summary_top_row + 9, col_offset + 18, wkend_profit_awaiting_formula, font = font_red_bold, border = cell_border['bottomleftright'], number_format=accounting_format)

        #mngment 
        # wk_sales_awaiting_formula = '=' + ' + '.join([f'SUM(F{cat_row_info[cat][0]}:F{cat_row_info[cat][1]})'for cat in cats_count_awaiting_pay_wk])
        # formatted_cell(ws, summary_top_row + 17, col_offset + 17, wk_sales_awaiting_formula, font = font_red_bold, border = cell_border['bottomleft'], number_format=accounting_format)
        # wkend_sales_awaiting_formula = '=' + ' + '.join([f'SUM(F{cat_row_info[cat][0]}:F{cat_row_info[cat][1]})'for cat in cats_count_awaiting_pay_wkend])
        # formatted_cell(ws, summary_top_row + 17, col_offset + 18, wkend_sales_awaiting_formula, font = font_red_bold, border = cell_border['bottomleftright'], number_format=accounting_format)
        
        # potential

        # wk_profit_potential_formula = '=' + ' + '.join([f'SUM(I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_for_potential_wk])
        # formatted_cell(ws, summary_top_row + 7, col_offset + 6, wk_profit_potential_formula, font = font_red, border = cell_border['topright'], number_format=accounting_format)
        # # wkend_profit_potential_formula = '=' + ' + '.join([f'SUM(I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_for_potential_wkend])
        # formatted_cell(ws, summary_top_row + 10, col_offset + 6, border = cell_border['right'])
        

        ws.conditional_formatting.add(
            f"I{cat_row_info['wk_complete_paid'][0]}:I{cat_row_info['ah_wo'][1]}", 
            FormulaRule(formula=[f'$K{cat_row_info["wk_complete_paid"][0]}="N"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )

        # 'wk_complete_paid',
        # 'wkend_complete_paid',

        for row in ws[f'W{50}:Y{curr_row-3}']:
            for cell in row:
                cell.border = cell_border_full


    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
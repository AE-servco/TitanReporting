import datetime as dt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule

import modules.helpers as helpers
import modules.lookup_tables as lookup



# function for each "box" in the summary
# function for the daily summary bits, one for monthly, one for weekly.
# function for actual jobs
#   - function for each job line maybe?
# calculate threshold function
#   - based on public holidays and days per week/month otherwise

class CommissionSpreadSheetExporter:
    def __init__(self, jobs_by_tech: dict[str, list[dict]], end_date: dt.date, timeframe: str, col_offset: int, scheme, holidays=[]):
        self.jobs_by_tech = jobs_by_tech
        self.end_date = end_date
        self.timeframe = timeframe
        self.curr_worksheet = None
        self.curr_row = 1
        self.bottom_row = 1
        self.job_count = 1
        self.col_offset = col_offset
        self.cat_row_info = {}
        self.threshold_day_num = 0
        self.curr_date = ""
        self.holidays = holidays
        self.scheme = scheme # "NSW" | "nonNSW"

        # lists of col letters for each day type
        self.weekend_cols = set()
        self.weekday_cols = set()

        # Keep track of rows with the profit dollar totals in them for summation cells later.
        self.rows_with_dollar_totals = set()
        self.rows_with_success_counts = set()
        self.rows_with_unsuccess_counts = set()

        # Important cells
        self.weekday_total_payable_row = None
        self.weekday_total_payable_col_num = None
        self.weekday_total_payable_col_letter = None
        
        self.weekend_total_payable_row = None
        self.weekend_total_payable_col_num = None
        self.weekend_total_payable_col_num_letter = None
        
        self.weekday_total_awaiting_payment_row = None
        self.weekday_total_awaiting_payment_col_num = None
        self.weekday_total_awaiting_payment_col_letetr = None
        
        self.weekend_total_awaiting_payment_row = None
        self.weekend_total_awaiting_payment_col_num = None
        self.weekend_total_awaiting_payment_col_letter = None

        self.review_count_total_row = None
        self.review_count_total_col_num = None
        self.review_count_total_col_letter = None

        self.cats_count_all = [
            'wk_complete_paid',
            'wkend_complete_paid',
            'wk_complete_unpaid',
            'wkend_complete_unpaid',
            'wk_wo',
            'wkend_wo',
            'ah_complete_paid',
            'ah_complete_unpaid',
            'ah_wo',
        ]

        self.cats_count_for_total = [
            'wk_complete_paid',
            'wkend_complete_paid',
            'wk_complete_unpaid',
            'wkend_complete_unpaid',
            'wk_wo',
            'wkend_wo',
        ]

        self.cats_count_for_afterhours_all = [
            'ah_complete_paid',
            'ah_complete_unpaid',
            'ah_wo',
        ]

        self.cats_count_for_emergency_paid = [
            # Not including weekend here because it is already coded in to use it elsewhere and it is friday arvo.
            'ah_complete_paid',
            'ph_complete_paid',
        ]

        self.cats_count_awaiting_pay_wk = [
            'wk_complete_unpaid',
            'wk_wo',
        ]

        self.cats_count_awaiting_pay_wkend = [
            'wkend_complete_unpaid',
            'wkend_wo',
            'ah_complete_unpaid',
            'ah_wo',
        ]

        self.cats_count_for_potential_wk = [
            'wk_complete_paid',
            'wk_complete_unpaid',
            'wk_wo',
            'wk_unsuccessful',
        ]

        self.cats_count_for_potential_wkend = [
            'wkend_complete_paid',
            'wkend_complete_unpaid',
            'wkend_wo',
            'wkend_unsuccessful',
            'ah_complete_paid',
            'ah_complete_unpaid',
            'ah_wo',
            'ah_unsuccessful',
        ]

        self.cats_count_for_unsuccessful = [
            'wk_unsuccessful',
            'wkend_unsuccessful',
            'ah_unsuccessful',
        ]

        # =========================== STYLING ===========================
        self.sales_color = '3da813'
        self.installer_color = 'e3b019'
        self.unknown_color = '8093f2'

        thin_border = Side(style='thin', color='000000') # Black thin border
        med_border = Side(style='medium', color='000000') # Black medium border
        thick_border = Side(style='thick', color='000000') # Black thick border
        double_border = Side(style='double', color='000000')

        self.yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
        self.blue_fill = PatternFill(patternType='solid', fgColor='3399F2')

        self.font_red = Font(color='FF0000')
        self.font_red_bold = Font(color='FF0000', bold=True)
        self.font_green_bold = Font(color='1F9C40', bold=True)
        self.font_green = Font(color='1F9C40')
        self.font_purple = Font(color='5F4F7D')
        self.font_bold = Font(bold=True)

        self.cell_border_full = Border(top=thin_border, bottom=thin_border, right=thin_border, left=thin_border)
        self.border_double_bottom = Border(bottom=double_border)

        self.accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        self.percentage_format = '0.00%'

        self.cell_border = {
            'full': Border(top=med_border, bottom=med_border, right=med_border, left=med_border),
            'topleft': Border(left=med_border, top=med_border),
            'topright': Border(right=med_border, top=med_border),
            'topleftright': Border(left=med_border, top=med_border, right=med_border),
            'bottomleft': Border(left=med_border, bottom=med_border),
            'bottomright': Border(right=med_border, bottom=med_border),
            'bottomtop': Border(top=med_border, bottom=med_border),
            'bottomtop_double': Border(top=med_border, bottom=double_border),
            'bottomtopright': Border(top=med_border, bottom=med_border, right=med_border),
            'bottomtopleft': Border(top=med_border, bottom=med_border, left=med_border),
            'bottomleftright': Border(left=med_border, bottom=med_border, right=med_border),
            'bottom': Border(bottom=med_border),
            'top': Border(top=med_border),
            'right': Border(right=med_border),
            'left': Border(left=med_border),
            'leftright': Border(left=med_border, right=med_border),
            'topthin': Border(top=thin_border),
        }

        self.align_center = Alignment(horizontal='center')

        # =========================== CONSTANTS ===========================
        self.CATEGORY_ORDER = {
            'wk_complete_paid': 'COMPLETED & PAID JOBS', 
            'wk_complete_unpaid': 'CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
            'wk_wo': 'CURRENT WORK ORDERS (AWAITING PAYMENT)', 
            'wk_unsuccessful': 'UNSUCCESSFUL JOBS',
            'wk_uncategorised': 'WEEK UNCATEGORISED',
            'wkend_complete_paid': 'WEEKEND COMPLETED & PAID JOBS', 
            'wkend_complete_unpaid': 'WEEKEND CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
            'wkend_wo': 'WEEKEND WORK ORDERS (AWAITING PAYMENT)', 
            'wkend_unsuccessful': 'WEEKEND UNSUCCESSFUL JOBS',
            'wkend_uncategorised': 'WEEKEND UNCATEGORISED',
            'ah_complete_paid': 'AFTERHOURS COMPLETED & PAID JOBS', 
            'ah_complete_unpaid': 'AFTERHOURS CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
            'ah_wo': 'AFTERHOURS WORK ORDERS (AWAITING PAYMENT)', 
            'ah_unsuccessful': 'AFTERHOURS UNSUCCESSFUL JOBS',
            'ah_uncategorised': 'AFTERHOURS UNCATEGORISED',
            'ph_complete_paid': 'PUBLIC HOLIDAYS COMPLETED & PAID JOBS', 
            'ph_complete_unpaid': 'PUBLIC HOLIDAYS CURRENT JOBS COMPLETED (AWAITING PAYMENT)', 
            'ph_wo': 'PUBLIC HOLIDAYS WORK ORDERS (AWAITING PAYMENT)', 
            'ph_unsuccessful': 'PUBLIC HOLIDAYS UNSUCCESSFUL JOBS',
            'ph_uncategorised': 'PUBLIC HOLIDAYS UNCATEGORISED',
            'prev': 'PREVIOUS JOBS COMPLETED & PAID (COMMISSION) - Modoras Team Please do ADD to THIS SECTION or AMEND/TOUCH', 
        }

    def _generate_sum_formula(self, cols: set, rows: set):
        """
        Generates an excel formula string that adds every combination of column and row from the given arguments. 
        E.g. '=A1+A2+A3+B1+B2+B3
        
        :param cols: Column letters
        :type cols: set
        :param rows: Row numbers
        :type rows: set
        """
        combinations = []
        for col in cols:
            for row in rows:
                combinations.append(f'{col}{row}')
        formula = '+'.join(combinations)
        return formula

    def formatted_cell(self, worksheet: Worksheet, row: int, col: int, val = None, font: Font | None=None, border: Border | None=None, number_format: str | None=None, fill: PatternFill | None=None):
        if val or val == 0:
            cell = worksheet.cell(row, col, val)
        else:
            cell = worksheet.cell(row, col)
        if font:
            cell.font = font
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format
        if fill:
            cell.fill = fill

        return cell

    def title_row(self, start_row: int = 1):
        ws = self.curr_worksheet
        col_offset = self.col_offset
        self.curr_row = start_row

        self.formatted_cell(ws, self.curr_row, col_offset + 1, 'WEEKLY COMMISSION', font = self.font_bold)
        self.formatted_cell(ws, self.curr_row, col_offset + 5, 'WEEK ENDING', font = self.font_red_bold)
        self.formatted_cell(ws, self.curr_row, col_offset + 6, self.end_date.strftime("%d/%m/%Y"), font = self.font_red_bold)
        return

    def job_count_box(self, start_row: int):
        ws = self.curr_worksheet
        col_offset = self.col_offset
        self.formatted_cell(ws, start_row, col_offset + 1, 'BOOKED JOBS',font = self.font_bold, border = self.cell_border['topleft'])

        # row and column definitions for this section
        label_col_num = col_offset + 1
        data_col_num = col_offset + 2
        label_col_letter = get_column_letter(label_col_num)
        data_col_letter = get_column_letter(data_col_num)

        success_count_row = start_row + 1
        unsuccess_count_row = start_row + 2
        success_rate_row = start_row + 3
        blank_row = start_row + 4
        avg_sale_row = start_row + 5

        self.formatted_cell(ws, success_count_row, label_col_num, 'SUCCESSFUL', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, unsuccess_count_row, label_col_num, 'UNSUCCESSFUL', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, success_rate_row, label_col_num, 'SUCCESSFUL (%)', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, avg_sale_row, label_col_num, 'AVERAGE SALE', font = self.font_bold, border = self.cell_border['bottomleft'])

        # Booked jobs
        self.formatted_cell(ws, start_row, data_col_num, f'={data_col_letter}{success_count_row}+{data_col_letter}{unsuccess_count_row}', font = self.font_bold, border = self.cell_border['topright'])
        # success rate
        self.formatted_cell(ws, success_rate_row, data_col_num, f'={data_col_letter}{success_count_row}/({data_col_letter}{success_count_row}+{data_col_letter}{unsuccess_count_row})', font = self.font_bold, border = self.cell_border['right'], number_format=self.percentage_format)
        
        if self.timeframe == 'monthly':
            # Success count (Only weekday at the moment. Does it need to be weekend?)
            self.formatted_cell(ws, success_count_row, data_col_num, f'={self._generate_sum_formula(self.weekday_cols, self.rows_with_success_counts)}', font = self.font_bold, border = self.cell_border['right'])
            # Unsuccess count
            self.formatted_cell(ws, unsuccess_count_row, data_col_num, f'={self._generate_sum_formula(self.weekday_cols, self.rows_with_unsuccess_counts)}', font = self.font_bold, border = self.cell_border['right'])
            # Avg sale
            self.formatted_cell(ws, avg_sale_row, data_col_num, f'={self.weekday_total_payable_col_letter}{self.weekday_total_payable_row}/{data_col_letter}{success_count_row}', font = self.font_bold, border = self.cell_border['bottomright'], number_format=self.accounting_format)
        elif self.timeframe == 'weekly':
            # TODO: REMOVE HARD CODED LETTERS
            # Success count (Only weekday at the moment. Does it need to be weekend?)
            self.formatted_cell(ws, success_count_row, data_col_num, '=P11', font = self.font_bold, border = self.cell_border['right'])
            # Unsuccess count
            self.formatted_cell(ws, unsuccess_count_row, data_col_num, '=P12', font = self.font_bold, border = self.cell_border['right'])
            # Avg sale
            self.formatted_cell(ws, avg_sale_row, data_col_num, '=P14', font = self.font_bold, border = self.cell_border['bottomright'], number_format=self.accounting_format)

        self.formatted_cell(ws, blank_row, label_col_num, border = self.cell_border['left'])
        self.formatted_cell(ws, blank_row, data_col_num, border = self.cell_border['right'])

        self.curr_row = avg_sale_row
        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row
        return

    def profit_target_box(self, start_row: int, tech_role: str):
        ws = self.curr_worksheet
        col_offset = self.col_offset

        self.formatted_cell(ws, start_row, col_offset + 4, 'PROFIT TARGET', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 6, self.threshold_day_num * 5000, font = self.font_bold, border = self.cell_border['topleft'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row, col_offset + 5, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 6, border = self.cell_border['topright'])
        self.formatted_cell(ws, start_row + 1, col_offset + 4, 'Tier 1', border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 1, col_offset + 5, f'<${self.threshold_day_num * 5000}')
        
        if self.scheme == "NSW":
            if tech_role == 'S':
                self.formatted_cell(ws, start_row + 1, col_offset + 6, '=0', border = self.cell_border['right'], number_format=self.percentage_format)
            else:
                self.formatted_cell(ws, start_row + 1, col_offset + 6, '=0.05', border = self.cell_border['right'], number_format=self.percentage_format)
        else:
            self.formatted_cell(ws, start_row + 1, col_offset + 6, '=0.05', border = self.cell_border['right'], number_format=self.percentage_format)
        
        self.formatted_cell(ws, start_row + 2, col_offset + 4, 'Tier 2', border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 2, col_offset + 5, f'>=${self.threshold_day_num * 5000}', border = self.cell_border['bottom'])
        if tech_role == "S":
            self.formatted_cell(ws, start_row + 2, col_offset + 6, '=0.1', border = self.cell_border['bottomright'], number_format=self.percentage_format)
        else:
            self.formatted_cell(ws, start_row + 2, col_offset + 6, '=0.05', border = self.cell_border['bottomright'], number_format=self.percentage_format)

        self.curr_row = start_row + 2
        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row
        return

    def payout_box(self, start_row: int):
        # TODO: make the letters dynamic in here
        ws = self.curr_worksheet
        col_offset = self.col_offset

        # if self.timeframe == 'weekly':
        #     totals_box_str = 'P10-S10'
        #     threshold_if_str = f'=IF(P10>={self.threshold_day_num * 5000},G5,G4)'
        #     emergency_str = "=Q10 + R10"
        # elif self.timeframe == 'monthly':
        totals_box_str = f'{self.weekday_total_payable_col_letter}{self.weekday_total_payable_row}-{self.weekday_total_awaiting_payment_col_letter}{self.weekday_total_awaiting_payment_row}'
        threshold_if_str = f'=IF({self.weekday_total_payable_col_letter}{self.weekday_total_payable_row}>={self.threshold_day_num * 5000},G5,G4)'

        # TODO: Add in afterhours and ph into this
        emergency_str = f'={self.weekend_total_payable_col_letter}{self.weekday_total_payable_row} - {self.weekend_total_awaiting_payment_col_letter}{self.weekend_total_awaiting_payment_row} + {self.ah_ph_total_col_letter}{self.ah_ph_total_row}'

        # cell coords for this box
        label_col_num = col_offset + 4
        actual_data_col_num = col_offset + 5
        potential_data_col_num = col_offset + 6
        super_data_col_num = col_offset + 7
        label_col_letter = get_column_letter(label_col_num)
        actual_data_col_letter = get_column_letter(actual_data_col_num)
        potential_data_col_letter = get_column_letter(potential_data_col_num)
        super_data_col_letter = get_column_letter(super_data_col_num)


        self.formatted_cell(ws, start_row, actual_data_col_num, 'ACTUAL', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, potential_data_col_num, 'POTENTIAL', font = self.font_bold, border = self.cell_border['topright'])
        self.formatted_cell(ws, start_row, super_data_col_num, 'Exc. SUPER', font = self.font_green_bold)
        self.formatted_cell(ws, start_row + 1, label_col_num, 'NET PROFIT', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row + 2, label_col_num, 'UNLOCKED', font = self.font_bold, border = self.cell_border['left'])

        self.formatted_cell(ws, start_row + 2, actual_data_col_num, threshold_if_str, border = self.cell_border['left'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, label_col_num, 'COMMISSION - PAY OUT', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 3, actual_data_col_num, f'={actual_data_col_letter}8*{actual_data_col_letter}9', border = self.cell_border['left'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 3, super_data_col_num, f'={actual_data_col_letter}10/1.12', font = self.font_green_bold, number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 4, label_col_num, 'EMERGENCY', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 4, actual_data_col_num, emergency_str, border = self.cell_border['left'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 5, label_col_num, 'EMERGENCY - PAY OUT', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 5, actual_data_col_num, f'={actual_data_col_letter}11*0.25', border = self.cell_border['left'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 5, super_data_col_num, f'={actual_data_col_letter}12/1.12', font = self.font_green_bold, number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 6, label_col_num, 'PREV. WEEK', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 6, actual_data_col_num, 0, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 7, label_col_num, 'PREV. WEEK - PAY OUT', font = self.font_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 7, actual_data_col_num, f'={actual_data_col_letter}13*0.05', border = self.cell_border['bottomleft'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 7, super_data_col_num, f'={actual_data_col_letter}14/1.12', font = self.font_green_bold, number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 8, label_col_num, '5 Star Review', font = self.font_green_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 8, actual_data_col_num, f'={actual_data_col_letter}16*50', border = self.cell_border['left'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 9, label_col_num, '5 Star Notes', font = self.font_green_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 9, actual_data_col_num, f'={self.review_count_total_col_letter}{self.review_count_total_row}', border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, potential_data_col_num, f'=IF({potential_data_col_letter}8>={self.threshold_day_num * 5000},{potential_data_col_letter}5,{potential_data_col_letter}4)', font=self.font_red, border = self.cell_border['right'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, potential_data_col_num, f'={potential_data_col_letter}8*{potential_data_col_letter}9', font = self.font_red, border = self.cell_border['right'], number_format=self.accounting_format)
        self.formatted_cell(ws, start_row + 5, potential_data_col_num, border = self.cell_border['right'])
        self.formatted_cell(ws, start_row + 6, potential_data_col_num, 0, font = self.font_red, border = self.cell_border['right'])
        self.formatted_cell(ws, start_row + 7, potential_data_col_num, f'=={potential_data_col_letter}13*0.05', font = self.font_red, border = self.cell_border['bottomright'], number_format=self.accounting_format)
        
        # Subtracting red from payout
        subtract_red_formula_wk = f'SUMIF(K{self.cat_row_info["wk_complete_paid"][0]}:K{self.cat_row_info["wk_complete_paid"][1]}, "N", I{self.cat_row_info["wk_complete_paid"][0]}:I{self.cat_row_info["wk_complete_paid"][1]})'
        
        self.formatted_cell(ws, start_row + 1, actual_data_col_num, f'={totals_box_str}-{subtract_red_formula_wk}', border = self.cell_border['topleft'], number_format=self.accounting_format)
        
        wk_profit_potential_formula = '=' + ' + '.join([f'SUM(I{self.cat_row_info[cat][0]}:I{self.cat_row_info[cat][1]})'for cat in self.cats_count_for_potential_wk])
        self.formatted_cell(ws, start_row + 1, potential_data_col_num, wk_profit_potential_formula, font = self.font_red, border = self.cell_border['topright'], number_format=self.accounting_format)
        # wkend_profit_potential_formula = '=' + ' + '.join([f'SUM(I{cat_row_info[cat][0]}:I{cat_row_info[cat][1]})'for cat in cats_count_for_potential_wkend])
        self.formatted_cell(ws, start_row + 4, potential_data_col_num, border = self.cell_border['right'])
        
        self.curr_row = start_row + 9
        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row
        return

    def doc_check_count_box(self, start_row: int):
        ws = self.curr_worksheet
        col_offset = self.col_offset

        self.formatted_cell(ws, start_row, col_offset + 10, 'PHOTOS', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 11, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 12, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 13, 'QUOTE', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 14, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 15, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 16, 'INVOICE', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 17, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 18, border = self.cell_border['topright'])
        self.formatted_cell(ws, start_row, col_offset + 19, border = self.cell_border['topright'])
        self.formatted_cell(ws, start_row + 1, col_offset + 10, 'BEFORE', font = self.font_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 1, col_offset + 11, 'AFTER', font = self.font_bold, border = self.cell_border['bottom'])
        self.formatted_cell(ws, start_row + 1, col_offset + 12, 'RECEIPT', font = self.font_bold, border = self.cell_border['bottom'])
        self.formatted_cell(ws, start_row + 1, col_offset + 13, 'DESCRIPTION', font = self.font_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 1, col_offset + 14, 'SIGNED', font = self.font_bold, border = self.cell_border['bottom'])
        self.formatted_cell(ws, start_row + 1, col_offset + 15, 'EMAILED', font = self.font_bold, border = self.cell_border['bottom'])
        self.formatted_cell(ws, start_row + 1, col_offset + 16, 'DESCRIPTION', font = self.font_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 1, col_offset + 17, 'SIGNED', font = self.font_bold, border = self.cell_border['bottom'])
        self.formatted_cell(ws, start_row + 1, col_offset + 18, 'EMAILED', font = self.font_bold, border = self.cell_border['bottom'])
        self.formatted_cell(ws, start_row + 1, col_offset + 19, '5 Star Review', font = self.font_bold, border = self.cell_border['bottomleftright'])
        self.formatted_cell(ws, start_row + 2, col_offset + 9, 'TAKEN', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row + 3, col_offset + 9, '%', font = self.font_bold, border = self.cell_border['bottomleft'])
        self.formatted_cell(ws, start_row + 3, col_offset + 19, border = self.cell_border['bottomleftright'])
        
        # TODO: can probably make this dynamic and a for loop
        end_of_comp_paid = self.cat_row_info['wk_complete_paid'][1]
        self.formatted_cell(ws, start_row + 2, col_offset + 10, f'=L{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 11, f'=M{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 12, f'=N{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 13, f'=O{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 14, f'=P{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 15, f'=Q{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 16, f'=R{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 17, f'=S{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        self.formatted_cell(ws, start_row + 2, col_offset + 18, f'=T{end_of_comp_paid+2}', font = self.font_bold, border = self.cell_border['left'])
        
        self.review_count_total_row = start_row + 2
        self.review_count_total_col_num = col_offset + 19
        self.review_count_total_col_letter = get_column_letter(self.review_count_total_col_num)

        review_count_formula = '=' + ' + '.join([f'SUM(U{self.cat_row_info[cat][0]}:U{self.cat_row_info[cat][1]})'for cat in self.cats_count_all])
        self.formatted_cell(ws, self.review_count_total_row, self.review_count_total_col_num, review_count_formula, font = self.font_bold, border = self.cell_border['leftright']) # total 5 star reviews

        self.formatted_cell(ws, start_row + 3, col_offset + 10, f'=L{end_of_comp_paid+2}/L{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 11, f'=M{end_of_comp_paid+2}/M{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 12, f'=N{end_of_comp_paid+2}/N{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 13, f'=O{end_of_comp_paid+2}/O{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 14, f'=P{end_of_comp_paid+2}/P{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 15, f'=Q{end_of_comp_paid+2}/Q{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 16, f'=R{end_of_comp_paid+2}/R{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 17, f'=S{end_of_comp_paid+2}/S{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        self.formatted_cell(ws, start_row + 3, col_offset + 18, f'=T{end_of_comp_paid+2}/T{end_of_comp_paid+1}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.percentage_format)
        
        
        self.curr_row = start_row + 3
        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row
        return

    def day_summaries_monthly(self, start_row: int):
        ws = self.curr_worksheet
        col_offset = self.col_offset

        dates_in_month = helpers.get_dates_in_month_datetime(self.end_date.year, self.end_date.month)

        # self.holidays = helpers.australian_public_holidays_between(self.end_date, dt.date(self.end_date.year, self.end_date.month, 1))

        self.threshold_day_num = helpers.get_threshold_days(dates_in_month, holidays=self.holidays)

        self.formatted_cell(ws, start_row, col_offset + 10, 'DAILY NET PROFIT', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 11, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 12, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 13, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 14, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 15, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 16, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row, col_offset + 12, 'WEEKDAY', font = self.font_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 13, 'WEEKEND', font = self.font_green_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 14, 'AH & PH', font = self.font_green_bold, border = self.cell_border['topleftright'])
        self.formatted_cell(ws, start_row + 1, col_offset + 10, 'Total Payable', font = self.font_bold, border = self.cell_border['top'])
        self.formatted_cell(ws, start_row + 2, col_offset + 10, 'Awaiting Payment', font = self.font_red_bold, border = self.cell_border['top'])

        profit_formulas = {day: '=' + ' + '.join([f'SUMIF(C{self.cat_row_info[cat][0]}:C{self.cat_row_info[cat][1]}, "{day.strftime("%d/%m/%Y")}", I{self.cat_row_info[cat][0]}:I{self.cat_row_info[cat][1]})'for cat in self.cats_count_for_total]) for day in dates_in_month}
        count_success_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{self.cat_row_info[cat][0]}:C{self.cat_row_info[cat][1]}, "{day.strftime("%d/%m/%Y")}")'for cat in self.cats_count_for_total]) for day in dates_in_month}
        count_unsuccess_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{self.cat_row_info[cat][0]}:C{self.cat_row_info[cat][1]}, "{day.strftime("%d/%m/%Y")}")'for cat in self.cats_count_for_unsuccessful]) for day in dates_in_month}

        day_start_row = start_row + 5
        reset_col = 10
        start_col = reset_col

        SUMMARY_COL_LENGTH = 0 # dynamic to start, but should be constant after first pass through. 
        for day in dates_in_month:
            day_row = day_start_row
            day_of_week = day.weekday()
            day_position = (day.day-1) % 14 # position when printing by fortnight, 0-indexed
            if day_position == 0:
                # If first in row, write the LHS words 
                self.formatted_cell(ws, day_row+2, col_offset+start_col-1, 'Successful')
                self.formatted_cell(ws, day_row+2+1, col_offset+start_col-1, 'Unsuccessful')
                self.formatted_cell(ws, day_row+2+2, col_offset+start_col-1, 'Success rate')
                self.formatted_cell(ws, day_row+2+3, col_offset+start_col-1, 'Avg sale')
            
            summary_font = self.font_bold

            curr_col = col_offset + start_col + day_position
            curr_col_letter = get_column_letter(curr_col)

            if day_of_week in [5,6]: # Weekend check
                summary_font = self.font_green_bold
                self.weekend_cols.add(curr_col_letter) 
            else:
                self.weekday_cols.add(curr_col_letter)

            # "Mon 1/11" row
            self.formatted_cell(ws, day_row, curr_col, day.strftime("%a %d/%m"), font = summary_font, border = self.cell_border['top'])
            day_row += 1

            # Profit row
            self.formatted_cell(ws, day_row, curr_col, profit_formulas[day], font = summary_font, number_format=self.accounting_format)
            self.rows_with_dollar_totals.add(day_row) # add this row to the list of rows with profit totals for summary generation later
            day_row += 1

            # Success count row
            self.formatted_cell(ws, day_row, curr_col, count_success_formulas[day])
            self.rows_with_success_counts.add(day_row) # add this row to the list of rows with success counts for summary generation later
            day_row += 1

            # Unsuccess count row
            self.formatted_cell(ws, day_row, curr_col, count_unsuccess_formulas[day])
            self.rows_with_unsuccess_counts.add(day_row) # add this row to the list of rows with unsuccess counts for summary generation later
            day_row += 1

            # Success rate row
            self.formatted_cell(ws, day_row, curr_col, f'={curr_col_letter}{day_row-2}/({curr_col_letter}{day_row-2}+{curr_col_letter}{day_row-1})', number_format=self.percentage_format)
            day_row += 1

            # Avg sale row
            self.formatted_cell(ws, day_row, curr_col, f'={curr_col_letter}{day_row-4}/{curr_col_letter}{day_row-3}', number_format=self.accounting_format)
            day_row += 1

            SUMMARY_COL_LENGTH = day_row - day_start_row
            if day_position == 13:
                # If last of fortnight, push row to next block
                day_start_row += SUMMARY_COL_LENGTH
        
        self.weekday_total_payable_row = start_row + 1
        self.weekday_total_payable_col_num = col_offset + 12
        self.weekday_total_payable_col_letter = get_column_letter(self.weekday_total_payable_col_num)
        
        self.weekend_total_payable_row = start_row + 1
        self.weekend_total_payable_col_num = col_offset + 13
        self.weekend_total_payable_col_letter = get_column_letter(self.weekend_total_payable_col_num)
        
        self.weekday_total_awaiting_payment_row = start_row + 2
        self.weekday_total_awaiting_payment_col_num = col_offset + 12
        self.weekday_total_awaiting_payment_col_letter = get_column_letter(self.weekday_total_awaiting_payment_col_num)
        
        self.weekend_total_awaiting_payment_row = start_row + 2
        self.weekend_total_awaiting_payment_col_num = col_offset + 13
        self.weekend_total_awaiting_payment_col_letter = get_column_letter(self.weekend_total_awaiting_payment_col_num)
        
        self.ah_ph_total_row = start_row + 1
        self.ah_ph_total_col_num = col_offset + 14
        self.ah_ph_total_col_letter = get_column_letter(self.weekend_total_awaiting_payment_col_num)

        self.formatted_cell(ws, self.weekday_total_payable_row, self.weekday_total_payable_col_num, f'={self._generate_sum_formula(self.weekday_cols, self.rows_with_dollar_totals)}', font = self.font_bold, border = self.cell_border['bottomleft'], number_format=self.accounting_format)
        self.formatted_cell(ws, self.weekend_total_payable_row, self.weekend_total_payable_col_num, f'={self._generate_sum_formula(self.weekend_cols, self.rows_with_dollar_totals)}', font = self.font_bold, border = self.cell_border['bottomright'], number_format=self.accounting_format)

        # Weekday awaiting payment total
        wk_profit_awaiting_formula = '=' + ' + '.join([f'SUM(I{self.cat_row_info[cat][0]}:I{self.cat_row_info[cat][1]})'for cat in self.cats_count_awaiting_pay_wk])
        self.formatted_cell(ws, self.weekday_total_awaiting_payment_row, self.weekday_total_awaiting_payment_col_num, wk_profit_awaiting_formula, font = self.font_red_bold, border = self.cell_border['bottomleft'], number_format=self.accounting_format)

        # Weekend awaiting payment total
        wkend_profit_awaiting_formula = '=' + ' + '.join([f'SUM(I{self.cat_row_info[cat][0]}:I{self.cat_row_info[cat][1]})'for cat in self.cats_count_awaiting_pay_wkend])
        self.formatted_cell(ws, self.weekend_total_awaiting_payment_row, self.weekend_total_awaiting_payment_col_num, wkend_profit_awaiting_formula, font = self.font_red_bold, border = self.cell_border['bottomleftright'], number_format=self.accounting_format)

        # Afterhour and Public Holiday total section
        total_ah_ph_formula = '=' + '+'.join([f'SUM(I{self.cat_row_info[cat][0]}:I{self.cat_row_info[cat][1]})' for cat in self.cats_count_for_emergency_paid])

        self.formatted_cell(ws, self.ah_ph_total_row, self.ah_ph_total_col_num, total_ah_ph_formula, font = self.font_green_bold, border = self.cell_border['bottomtop'], number_format=self.accounting_format) 

        return

    def day_summaries_weekly(self, start_row: int, management: bool = False):
#         =SUMIF(B{cat_start}:B{cat_end}, date_str, E{cat_start}:E{cat_end}) -- sales
#         =SUMIF(B{cat_start}:B{cat_end}, date_str, H{cat_start}:H{cat_end}) -- profit

        ws = self.curr_worksheet
        col_offset = self.col_offset

        # TODO: Make this more robust, if a Sunday is not selected in the filter, this is all wrong...
        dates = {
            'monday': self.end_date - dt.timedelta(days=6),
            'tuesday': self.end_date - dt.timedelta(days=5),
            'wednesday': self.end_date - dt.timedelta(days=4),
            'thursday': self.end_date - dt.timedelta(days=3),
            'friday': self.end_date - dt.timedelta(days=2),
            'saturday': self.end_date - dt.timedelta(days=1),
            'sunday': self.end_date
        }

        # self.holidays = helpers.australian_public_holidays_between(self.end_date, dates['monday'])

        self.threshold_day_num = helpers.get_threshold_days(list(dates.values()), holidays=self.holidays)
    
        date_strs = {day: date.strftime("%d/%m/%Y") for day, date in dates.items()}

        column_to_sum = "F" if management else "I"

        if management:
            # Sales section for management summary if needed
            self.formatted_cell(ws, start_row, col_offset + 10, 'Management Summary - SALES ONLY - COLUMN F', font = self.font_bold, border = self.cell_border['topleft'], fill=self.blue_fill)
            for i in range(9):
                self.formatted_cell(ws, start_row, col_offset + 11 + i, border = self.cell_border['top'], fill=self.blue_fill)

        else:
            self.formatted_cell(ws, start_row, col_offset + 10, 'DAILY NET PROFIT', font = self.font_bold, border = self.cell_border['topleft'])
            for i in range(9):
                self.formatted_cell(ws, start_row, col_offset + 11 + i, border = self.cell_border['top'])

        self.formatted_cell(ws, start_row, col_offset + 20, 'WEEKDAY', font = self.font_red_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row, col_offset + 21, 'WEEKEND', font = self.font_red_bold, border = self.cell_border['topleftright'])
        self.formatted_cell(ws, start_row + 1, col_offset + 10, 'MON', font = self.font_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 11, 'TUE', font = self.font_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 12, 'WED', font = self.font_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 13, 'THU', font = self.font_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 14, 'FRI', font = self.font_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 15, 'TOTAL', font = self.font_bold, border = self.cell_border['bottomtop_double'])

        self.formatted_cell(ws, start_row + 1, col_offset + 16, 'SAT', font = self.font_green_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 17, 'SUN', font = self.font_green_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 18, 'TOTAL (WKEND)', font = self.font_green_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 19, 'TOTAL (AH & PH)', font = self.font_green_bold, border = self.cell_border['bottomtop_double'])
        self.formatted_cell(ws, start_row + 1, col_offset + 20, 'Awaiting Payment', font = self.font_red_bold, border = self.cell_border['topleft'])
        self.formatted_cell(ws, start_row + 1, col_offset + 21, 'Awaiting Payment', font = self.font_red_bold, border = self.cell_border['topleftright'])

        self.formatted_cell(ws, start_row + 2, col_offset + 9, border=self.cell_border['right'])
        self.formatted_cell(ws, start_row + 3, col_offset + 9, 'Successful')
        self.formatted_cell(ws, start_row + 4, col_offset + 9, 'Unsuccessful')
        self.formatted_cell(ws, start_row + 5, col_offset + 9, 'Success rate')
        self.formatted_cell(ws, start_row + 6, col_offset + 9, 'Avg sale')

        days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']

        # Daily totals
        daily_formulas = {day: '=' + ' + '.join([f'SUMIF(C{self.cat_row_info[cat][0]}:C{self.cat_row_info[cat][1]}, "{date_strs[day]}", {column_to_sum}{self.cat_row_info[cat][0]}:{column_to_sum}{self.cat_row_info[cat][1]})'for cat in self.cats_count_for_total]) for day in days}
        # Daily success counts
        count_success_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{self.cat_row_info[cat][0]}:C{self.cat_row_info[cat][1]}, "{date_strs[day]}")'for cat in self.cats_count_for_total]) for day in days}
        # Daily unsuccessful counts
        count_unsuccess_formulas = {day: '=' + ' + '.join([f'COUNTIF(C{self.cat_row_info[cat][0]}:C{self.cat_row_info[cat][1]}, "{date_strs[day]}")'for cat in self.cats_count_for_unsuccessful]) for day in days}
        
        for idx, dayname in enumerate(days):
            if dayname in ['saturday', 'sunday']:
                # +1 to compensate for total column in between weekday and weekends.
                self.formatted_cell(ws, start_row + 2, col_offset + 10 + idx + 1, daily_formulas[dayname], font = self.font_bold, border = self.cell_border['bottom'], number_format=self.accounting_format)
                self.formatted_cell(ws, start_row + 3, col_offset + 10 + idx + 1, count_success_formulas[dayname]) 
                self.formatted_cell(ws, start_row + 4, col_offset + 10 + idx + 1, count_unsuccess_formulas[dayname])
            else:
                self.formatted_cell(ws, start_row + 2, col_offset + 10 + idx, daily_formulas[dayname], font = self.font_bold, border = self.cell_border['bottom'], number_format=self.accounting_format)
                self.formatted_cell(ws, start_row + 3, col_offset + 10 + idx, count_success_formulas[dayname])
                self.formatted_cell(ws, start_row + 4, col_offset + 10 + idx, count_unsuccess_formulas[dayname])

        # Weekday total section
        total_wk_formula = '=' + '+'.join([f'SUM({get_column_letter(col_offset + 10 + i)}{start_row + 2})' for i in range(5)])
        weekday_total_col_num = col_offset + 15
        weekday_total_col_letter = get_column_letter(weekday_total_col_num)
        weekday_total_row = start_row + 2
        if not management:
            # Only set these if not management summary
            self.weekday_total_payable_col_num = weekday_total_col_num
            self.weekday_total_payable_col_letter = weekday_total_col_letter
            self.weekday_total_payable_row = weekday_total_row
        self.formatted_cell(ws, weekday_total_row, weekday_total_col_num, total_wk_formula, font = self.font_bold, border = self.cell_border['bottomright'], number_format=self.accounting_format)
        
        # Weekend total section
        total_wkend_formula = '=' + '+'.join([f'SUM({get_column_letter(col_offset + 10 + i)}{start_row + 2})' for i in [6,7]]) # 6,7 because total column in between weekdays and weekend
        weekend_total_col_num = col_offset + 18
        weekend_total_col_letter = get_column_letter(weekend_total_col_num)
        weekend_total_row = start_row + 2
        if not management:
            # Only set if not management summary
            self.weekend_total_payable_col_num = weekend_total_col_num
            self.weekend_total_payable_col_letter = weekend_total_col_letter
            self.weekend_total_payable_row = weekend_total_row
        self.formatted_cell(ws, weekend_total_row, weekend_total_col_num, total_wkend_formula, font = self.font_green_bold, border = self.cell_border['bottomtop'], number_format=self.accounting_format) 
        
        # Afterhour and Public Holiday total section
        # subtract_red_ah_ph_formula = '(' + '+'.join([f'SUMIF(K{self.cat_row_info[cat][0]}:K{self.cat_row_info[cat][1]}, "N", {column_to_sum}{self.cat_row_info[cat][0]}:{column_to_sum}{self.cat_row_info[cat][1]})' for cat in self.cats_count_for_emergency_paid]) + ')'

        total_ah_ph_formula = '=' + '+'.join(
            [f'SUM({column_to_sum}{self.cat_row_info[cat][0]}:{column_to_sum}{self.cat_row_info[cat][1]}) - SUMIF(K{self.cat_row_info[cat][0]}:K{self.cat_row_info[cat][1]}, "N", {column_to_sum}{self.cat_row_info[cat][0]}:{column_to_sum}{self.cat_row_info[cat][1]})' for cat in self.cats_count_for_emergency_paid])
        # total_ah_ph_formula = '=' + '+'.join([f'SUM({get_column_letter(col_offset + 10 + i)}{start_row + 2})' for i in [6,7]]) # 6,7 because total column in between weekdays and weekend
        ah_ph_total_col_num = col_offset + 19
        ah_ph_total_col_letter = get_column_letter(ah_ph_total_col_num)
        ah_ph_total_row = start_row + 2
        if not management:
            # Only set if not management summary
            self.ah_ph_total_col_num = ah_ph_total_col_num
            self.ah_ph_total_col_letter = ah_ph_total_col_letter
            self.ah_ph_total_row = ah_ph_total_row
        self.formatted_cell(ws, ah_ph_total_row, ah_ph_total_col_num, total_ah_ph_formula, font = self.font_green_bold, border = self.cell_border['bottomtop'], number_format=self.accounting_format) 

        # Weekday awaiting payment section
        awaiting_pay_total_wk = '=' + ' + '.join([f'SUM({column_to_sum}{self.cat_row_info[cat][0]}:{column_to_sum}{self.cat_row_info[cat][1]})'for cat in self.cats_count_awaiting_pay_wk])
        weekday_total_awaiting_col_num = col_offset + 20
        weekday_total_awaiting_col_letter = get_column_letter(weekday_total_awaiting_col_num)
        weekday_total_awaiting_row = start_row + 2
        if not management:
            # Only set if not management summary
            self.weekday_total_awaiting_payment_col_num = weekday_total_awaiting_col_num
            self.weekday_total_awaiting_payment_col_letter = weekday_total_awaiting_col_letter
            self.weekday_total_awaiting_payment_row = weekday_total_awaiting_row
        self.formatted_cell(ws, weekday_total_awaiting_row, weekday_total_awaiting_col_num, awaiting_pay_total_wk, font = self.font_red_bold, border = self.cell_border['full'], number_format=self.accounting_format)

        # Weekend awaiting payment section
        awaiting_pay_total_wkend = '=' + ' + '.join([f'SUM({column_to_sum}{self.cat_row_info[cat][0]}:{column_to_sum}{self.cat_row_info[cat][1]})'for cat in self.cats_count_awaiting_pay_wkend])
        weekend_total_awaiting_col_num = col_offset + 21
        weekend_total_awaiting_col_letter = get_column_letter(weekend_total_awaiting_col_num)
        weekend_total_awaiting_row = start_row + 2
        if not management:
            # Only set if not management summary
            self.weekend_total_awaiting_payment_col_num = weekend_total_awaiting_col_num
            self.weekend_total_awaiting_payment_col_letter = weekend_total_awaiting_col_letter
            self.weekend_total_awaiting_payment_row = weekend_total_awaiting_row
        self.formatted_cell(ws, weekend_total_awaiting_row, weekend_total_awaiting_col_num, awaiting_pay_total_wkend, font = self.font_red_bold, border = self.cell_border['full'], number_format=self.accounting_format)


        # Total success counts weekday (dont need weekend)
        total_wk_count_success_formula = '=' + '+'.join([f'SUM({get_column_letter(col_offset + 10 + i)}{start_row + 3})' for i in range(5)])
        self.formatted_cell(ws, start_row + 3, col_offset + 15, total_wk_count_success_formula)

        # Total unsuccessful counts weekday (don't need weekend)
        total_wk_count_unsuccess_formula = '=' + '+'.join([f'SUM({get_column_letter(col_offset + 10 + i)}{start_row + 4})' for i in range(5)])
        self.formatted_cell(ws, start_row + 4, col_offset + 15, total_wk_count_unsuccess_formula)

        # Success rate and Avg sale boxes
        curr_col = col_offset + 10
        for i in range(8):
            curr_col_letter = get_column_letter(curr_col + i)
            # success rate
            self.formatted_cell(ws, start_row + 5, curr_col + i, f'={curr_col_letter}{start_row + 5-2}/({curr_col_letter}{start_row + 5-2}+{curr_col_letter}{start_row + 5-1})', number_format=self.percentage_format)
            # avg sale
            self.formatted_cell(ws, start_row + 6, curr_col + i, f'={curr_col_letter}{start_row + 5-3}/{curr_col_letter}{start_row + 5-2}', number_format=self.accounting_format)

        return

    def job_section_title(self, row):
        ws = self.curr_worksheet
        col_offset = self.col_offset

        ws.cell(row, col_offset + 1, 'JOB DETAILS').border = self.cell_border['topleft']
        ws.cell(row, col_offset + 2).border = self.cell_border['top']
        ws.cell(row, col_offset + 3).border = self.cell_border['top']
        ws.cell(row, col_offset + 4).border = self.cell_border['top']
        ws.cell(row, col_offset + 5, 'JOB AMOUNT').border = self.cell_border['topleft']
        ws.cell(row, col_offset + 6).border = self.cell_border['top']
        ws.cell(row, col_offset + 7).border = self.cell_border['top']
        ws.cell(row, col_offset + 8).border = self.cell_border['top']
        ws.cell(row, col_offset + 9).border = self.cell_border['top']
        ws.cell(row, col_offset + 10).border = self.cell_border['topleft']
        ws.cell(row, col_offset + 11, 'PHOTOS').border = self.cell_border['topleft']
        ws.cell(row, col_offset + 12).border = self.cell_border['top']
        ws.cell(row, col_offset + 13).border = self.cell_border['top']
        ws.cell(row, col_offset + 14, 'QUOTE').border = self.cell_border['topleft']
        ws.cell(row, col_offset + 15).border = self.cell_border['top']
        ws.cell(row, col_offset + 16).border = self.cell_border['top']
        ws.cell(row, col_offset + 17, 'INVOICE').border = self.cell_border['topleft']
        ws.cell(row, col_offset + 18).border = self.cell_border['top']
        ws.cell(row, col_offset + 19).border = self.cell_border['top']
        ws.cell(row, col_offset + 20).border = self.cell_border['top']
        ws.cell(row, col_offset + 21, 'NOTES').border = self.cell_border['topright']

        row += 1
        self.curr_row = row

        ws.cell(row, col_offset + 1, 'JOB STATUS').border = self.cell_border['bottomleft']
        ws.cell(row, col_offset + 2, 'DATE').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 3, 'JOB #').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 4, 'SUBURB').border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 5, 'AMOUNT EXC. GST').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 6, 'MATERIALS').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 7, 'MERCHANT FEES').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 8, 'NET PROFIT').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 9, 'PAID').border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 10, 'DOC CHECK DONE').border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 11, 'BEFORE').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 12, 'AFTER').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 13, 'RECEIPT').border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 14, 'DESCRIPTION').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 15, 'SIGNED').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 16, 'EMAILED').border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 17, 'DESCRIPTION').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 18, 'SIGNED').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 19, 'EMAILED').border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 20, '5 Star Review').border = self.cell_border['bottom']
        ws.cell(row, col_offset + 21).border = self.cell_border['bottomright']
        ws.cell(row, col_offset + 22, 'EFTPOS').border = self.cell_border['bottomtop']
        ws.cell(row, col_offset + 23, 'CASH').border = self.cell_border['bottomtop']
        ws.cell(row, col_offset + 24, 'Payment Plan').border = self.cell_border['bottomtopright']
        
        self.curr_row += 1
        return
    
    def put_job_row(self, job: dict, row: int, cat_font: Font):
        ws = self.curr_worksheet
        col_offset = self.col_offset

        if job['first_appt_start_str'] != self.curr_date:
            job_border = self.cell_border['topthin']
        else:
            job_border = None
        self.curr_date = job['first_appt_start_str']

        if job['complaint_tag_present']:
            self.formatted_cell(ws, row, col_offset, 'COMPLAINT', font=cat_font)#, border = job_border)
        self.formatted_cell(ws, row, col_offset + 1, self.job_count, font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 2, job['first_appt_start_str'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 3, int(job['num']), font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 4, job['suburb'], font=cat_font, border = job_border)
        if not job['unsuccessful']:
            self.formatted_cell(ws, row, col_offset + 5, job['inv_subtotal'], font=cat_font, border = job_border, number_format=self.accounting_format)
            self.formatted_cell(ws, row, col_offset + 6, job['inv_subtotal']*0.2, font=cat_font, border = job_border, number_format=self.accounting_format)
            self.formatted_cell(ws, row, col_offset + 6, font=cat_font, border = job_border).comment = Comment(job['summary'], "automation")
        else:
            self.formatted_cell(ws, row, col_offset + 5, job['open_est_subtotal'], font=cat_font, border = job_border, number_format=self.accounting_format)
            self.formatted_cell(ws, row, col_offset + 6, "", font=cat_font, border = job_border)
        # 7
        self.formatted_cell(ws, row, col_offset + 7, "", font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 8, f"={get_column_letter(col_offset + 5)}{row}-{get_column_letter(col_offset + 6)}{row}-{get_column_letter(col_offset + 7)}{row}", font=cat_font, border = job_border, number_format=self.accounting_format)
        self.formatted_cell(ws, row, col_offset + 9, job.get('payment_types', ''), font=cat_font, border = job_border)
        # 10 TODO: all doc checks complete
        doc_check_complete_col = f'=IF(OR({", ".join([f"{get_column_letter(col_offset + i)}{row}=0" for i in range(11,20)])}), "N","Y")'
        self.formatted_cell(ws, row, col_offset + 10, doc_check_complete_col, font=cat_font, border = job_border).alignment = self.align_center
        self.formatted_cell(ws, row, col_offset + 11, job['Before Photo'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 12, job['After Photo'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 13, job['Receipt Photo'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 14, job['Quote Description'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 15, job['Quote Signed'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 16, job['Quote Emailed'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 17, job['Invoice Description'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 18, job['Invoice Signed'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 19, job['Invoice Emailed'], font=cat_font, border = job_border)
        self.formatted_cell(ws, row, col_offset + 20, job['5 Star Review'], font=cat_font, border = job_border)

        self.formatted_cell(ws, row, col_offset + 25, f"=ROUND(F{row}*1.1,2)", font=self.font_purple)
        self.formatted_cell(ws, row, col_offset + 26, f"=ROUND(Z{row} - W{row} - X{row} - Y{row},2)", font=self.font_purple)
        
        payments = job.get('payment_details')

        if type(payments) == str:
            payments = payments.split(', ')
        
            all_payment_types = lookup.get_all_payment_types()
            p_types = {p_type: [] for p_type in set(all_payment_types.values())}
            
            for p in payments:
                p_type, amount = p.split('|')
                p_types[p_type].append(amount)
                
            if p_types['EFT'] or p_types['CC']:
                self.formatted_cell(ws, row, col_offset + 22, f"={'+'.join(p_types['EFT'] + p_types['CC'])}", font=cat_font, number_format=self.accounting_format)
            if p_types['Cash']:
                self.formatted_cell(ws, row, col_offset + 23, f"={'+'.join(p_types['Cash'])}", font=cat_font, number_format=self.accounting_format)
            if p_types['PP']:
                self.formatted_cell(ws, row, col_offset + 24, f"={'+'.join(p_types['PP'])}", font=cat_font, number_format=self.accounting_format)
        return
    
    def category_totals_row(self, cat: str, amt_col: int, materials_col: int, merchantf_col: int, profit_col: int):
        # totals row
        ws = self.curr_worksheet
        col_offset = self.col_offset
        curr_row = self.curr_row
        cat_row_start, cat_row_end = self.cat_row_info[cat]

        amt_letter = get_column_letter(col_offset + amt_col)
        materials_letter = get_column_letter(col_offset + materials_col)
        merchantf_letter = get_column_letter(col_offset + merchantf_col)
        profit_letter = get_column_letter(col_offset + profit_col)
        ws.cell(curr_row, col_offset + amt_col, f"=SUM({amt_letter}{cat_row_start}:{amt_letter}{cat_row_end})").number_format = self.accounting_format
        ws.cell(curr_row, col_offset + materials_col, f"=SUM({materials_letter}{cat_row_start}:{materials_letter}{cat_row_end})").number_format = self.accounting_format
        ws.cell(curr_row, col_offset + merchantf_col, f"=SUM({merchantf_letter}{cat_row_start}:{merchantf_letter}{cat_row_end})").number_format = self.accounting_format
        ws.cell(curr_row, col_offset + profit_col, f"=SUM({profit_letter}{cat_row_start}:{profit_letter}{cat_row_end})").number_format = self.accounting_format
        ws.cell(curr_row, col_offset + amt_col).border = self.border_double_bottom
        ws.cell(curr_row, col_offset + materials_col).border = self.border_double_bottom
        ws.cell(curr_row, col_offset + merchantf_col).border = self.border_double_bottom
        ws.cell(curr_row, col_offset + profit_col).border = self.border_double_bottom

        if cat == 'wk_complete_paid':
            for i in range(10):
                curr_col = col_offset + 11 + i
                curr_col_letter = get_column_letter(curr_col)
                self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({curr_col_letter}{cat_row_start}:{curr_col_letter}{cat_row_end})")
                self.formatted_cell(ws, curr_row + 1, curr_col, f"=SUM({curr_col_letter}{cat_row_start}:{curr_col_letter}{cat_row_end})")
            self.curr_row += 1

        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # curr_col += 1
            # self.formatted_cell(ws, curr_row, curr_col, f"=COUNT({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")


            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
            # self.formatted_cell(ws, curr_row+1, curr_col, f"=SUM({get_column_letter(curr_col)}{cat_row_start}:{get_column_letter(curr_col)}{cat_row_end})")
        return
    
    def put_job_category(self, cat: str, cat_text: str, job_cats:dict, start_row: int):
        ws = self.curr_worksheet
        col_offset = self.col_offset
        cat_font = None
        
        if cat in ['wk_complete_unpaid', 'wk_wo', 'wkend_complete_unpaid', 'wkend_wo']:
            cat_font = self.font_green
        self.formatted_cell(ws, start_row, col_offset + 1,cat_text)#, font=cat_font)
        cat_row = start_row
        cat_row += 1

        if cat == 'prev':
            for row in ws[f'B{cat_row-1}:X{cat_row+7}']: # TODO: Make these column letters dynamic if possible?
                for cell in row:
                    cell.fill = self.yellow_fill
            for i in range(1,9):
                ws.cell(cat_row, col_offset + 1, i)
                cat_row += 1

        else:
            jobs = job_cats.get(cat, [])
            if not jobs:
                cat_row += 1
                # self.curr_row = cat_row
            self.job_count = 1
            for job in jobs:
                self.put_job_row(job, cat_row, cat_font)
                cat_row += 1
                self.job_count += 1
        
        self.curr_row = cat_row # Don't need to mess around with curr_row this much, just making sure it works for now.
        # if self.bottom_row < self.curr_row:
        #     self.bottom_row = self.curr_row
        
        # add 5 rows for any extras to add
        self.curr_row += 5
        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row

        # Fill in info 
        self.cat_row_info[cat] = (start_row + 1, self.curr_row - 1)

        self.category_totals_row(cat, amt_col=5, materials_col=6, merchantf_col=7, profit_col=8)
        self.curr_row += 2
        if self.bottom_row < self.curr_row:
            self.bottom_row = self.curr_row
        return

    def extra_formatting(self):
        ws = self.curr_worksheet

        # ----------------------------------------------------------------------------------------
        # Doc check formatting 0s to red
        dc_start_col_letter = get_column_letter(self.col_offset + 10)
        dc_end_col_letter = get_column_letter(self.col_offset + 20)
        for cat in self.cat_row_info.keys():
            cat_row_start, cat_row_end = self.cat_row_info[cat]
            dc_range = f'{dc_start_col_letter}{cat_row_start}:{dc_end_col_letter}{cat_row_end}'
            ws.conditional_formatting.add(
                dc_range, 
                CellIsRule(operator="equal", formula=["0"], font=self.font_red)
            )
        # ----------------------------------------------------------------------------------------
        # Highlighting red for no doc check
        ws.conditional_formatting.add(
            f"I{self.cat_row_info['wk_complete_paid'][0]}:I{self.cat_row_info['ah_wo'][1]}", 
            FormulaRule(formula=[f'$K{self.cat_row_info["wk_complete_paid"][0]}="N"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )
        # ----------------------------------------------------------------------------------------
        # Border on RHS things
        for row in ws[f'W{self.job_start_row + 2}:Y{self.bottom_row-3}']:
            for cell in row:
                cell.border = self.cell_border_full
        # ----------------------------------------------------------------------------------------

    def build_sheet(self, wb: Workbook, tech: str):
        job_cats = self.jobs_by_tech[tech]
        if not self.curr_worksheet:
            self.curr_worksheet = wb.active
            self.curr_worksheet.title = tech[:-1]
            self.first_sheet_created = True
        else:
            self.curr_worksheet = wb.create_sheet(title=tech[:-1])

        if tech[-1] == 'S':
            self.curr_worksheet.sheet_properties.tabColor = self.sales_color
        elif tech[-1] == 'I':
            self.curr_worksheet.sheet_properties.tabColor = self.installer_color
        else:
            self.curr_worksheet.sheet_properties.tabColor = self.unknown_color
        
        # ========================== BOTTOM SECTION - JOBS ==========================
        if self.timeframe == 'monthly':
            self.job_start_row = 50
        elif self.timeframe == 'weekly':
            self.job_start_row = 25
        # self.curr_row = job_start_row
        self.job_section_title(self.job_start_row)

        for cat, cat_text in self.CATEGORY_ORDER.items():
            self.put_job_category(cat, cat_text, job_cats, self.curr_row)
        
        self.extra_formatting()

        # ========================= TOP SECTION - SUMMARIES =========================
        # Will need to call some of these after doing jobs because they will need to know cat start and ends etc.
        self.title_row(start_row=1)
        if self.timeframe == 'weekly':
            self.day_summaries_weekly(start_row=8, management=False)
            self.day_summaries_weekly(start_row=16, management=True)
        elif self.timeframe == 'monthly':
            self.day_summaries_monthly(start_row=8)
        else:
            raise ValueError("Timeframe must be 'weekly' or 'monthly'")
        self.job_count_box(start_row=3)
        self.payout_box(start_row=7)
        self.profit_target_box(start_row=3, tech_role=tech[-1])
        self.doc_check_count_box(start_row=2)

    def build_workbook(self):
        # Final function that actually builds everything
        wb = Workbook()

        for tech in sorted(self.jobs_by_tech.keys()):
            self.build_sheet(wb, tech)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()

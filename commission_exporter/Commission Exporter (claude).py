import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Tuple
from servicetitan_api_client import ServiceTitanClient
import modules.google_store as gs
import modules.helpers as helpers
import modules.templates as templates


class CommissionCalculator:
    def __init__(self, st_client):
        self.client: ServiceTitanClient = st_client
        
    def fetch_technicians(self) -> List[Dict]:
        """Fetch all technicians from ServiceTitan"""
        url = self.client.build_url('settings', 'technicians')
        response = self.client.get_all(url)
        return response.get('data', [])
    
    def fetch_jobs_for_technician(self, tech_id: int, start_date: str, end_date: str) -> List[Dict]:
        # TODO check date formatting (look in main)
        """Fetch all jobs for a technician within date range"""
        params = {
            'createdOnOrAfter': start_date,
            'createdBefore': end_date,
            'technicianId': tech_id
        }
        url = self.client.build_url('jpm', 'jobs')
        response = self.client.get_all(url, params=params)
        return response.get('data', [])
    
    def check_doc_checks(self, external_data: Dict) -> bool:
        """Check if all doc checks pass from externalData"""
        if not external_data:
            return False
        
        # Add your specific doc check fields here
        required_checks = [
            'photos_before',
            'photos_after', 
            'photos_receipt',
            'quote_description',
            'quote_signed',
            'quote_emailed',
            'invoice_description',
            'invoice_signed',
            'invoice_emailed'
        ]
        
        # Check all required fields are present and True/Yes
        for check in required_checks:
            value = external_data.get(check)
            if not value or str(value).lower() not in ['true', 'yes', '1']:
                return False
        
        return True
    
    def categorize_job(self, job: Dict) -> Tuple[str, Dict]:
        """Categorize job and extract relevant information"""
        # Extract basic info
        job_info = {
            'date': datetime.fromisoformat(job['createdOn'].replace('Z', '+00:00')).strftime('%d/%m/%Y'),
            'job_number': job.get('number', ''),
            'job_id': job.get('id', ''),
            'location': job.get('location', {}).get('address', {}).get('city', ''),
            'suburb': job.get('location', {}).get('address', {}).get('city', ''),
            'amount_exc_gst': job.get('summary', {}).get('subtotal', 0),
            'materials': 0,  # To be filled manually
            'merchant_fees': 0,
            'net_profit': 0,  # Will calculate after materials input
            'doc_checks': self.check_doc_checks(job.get('externalData', {})),
            'invoice_paid': job.get('invoice', {}).get('paid', False),
            'job_status': job.get('jobStatus', ''),
            'business_unit': job.get('businessUnit', {}).get('name', ''),
        }
        
        # Calculate net profit (will be refined with manual materials input)
        job_info['net_profit'] = job_info['amount_exc_gst'] - job_info['materials']
        
        # Categorize job
        if job_info['job_status'] in ['Completed'] and job_info['invoice_paid']:
            category = 'completed_paid'
        elif job_info['job_status'] in ['Completed']:
            category = 'awaiting_payment'
        elif job_info['job_status'] in ['Canceled', 'Cancelled'] or not job.get('sold', True):
            category = 'unsuccessful'
        else:
            category = 'unsuccessful'
            
        return category, job_info
    
    def calculate_commission(self, total_net_profit: float, scheme: str, doc_checks_pass: bool) -> float:
        """Calculate commission based on scheme and doc checks"""
        if not doc_checks_pass:
            return 0
        
        threshold = 25000
        
        if scheme == 'scheme1':
            # All or nothing: 10% only if threshold reached
            return total_net_profit * 0.10 if total_net_profit >= threshold else 0
        else:  # scheme2
            # 5% before threshold, 10% after threshold is reached
            return total_net_profit * 0.10 if total_net_profit >= threshold else total_net_profit * 0.05
    
    def process_jobs(self, jobs: List[Dict]) -> Dict:
        """Process all jobs and categorize them"""
        categorized = {
            'completed_paid': [],
            'completed_paid_weekend': [],
            'awaiting_payment': [],
            'awaiting_payment_weekend': [],
            'unsuccessful': [],
            'unsuccessful_weekend': []
        }
        
        for job in jobs:
            category, job_info = self.categorize_job(job)
            
            # Check if weekend (Saturday=5, Sunday=6)
            job_date = datetime.fromisoformat(job['createdOn'].replace('Z', '+00:00'))
            is_weekend = job_date.weekday() >= 5
            
            if is_weekend:
                categorized[f'{category}_weekend'].append(job_info)
            else:
                categorized[category].append(job_info)
        
        return categorized
    
    def generate_excel(self, tech_name: str, categorized_jobs: Dict, 
                       start_date: str, end_date: str, scheme: str, 
                       materials_input: Dict = None) -> BytesIO:
        """Generate Excel file matching the template"""
        
        # Update materials if provided
        if materials_input:
            for category in categorized_jobs.values():
                for job in category:
                    job_key = f"{job['job_number']}"
                    if job_key in materials_input:
                        job['materials'] = materials_input[job_key]
                        job['net_profit'] = job['amount_exc_gst'] - job['materials']
        
        # Calculate totals
        total_net_profit = sum(job['net_profit'] for job in categorized_jobs['completed_paid'])
        total_amount = sum(job['amount_exc_gst'] for job in categorized_jobs['completed_paid'])
        successful_count = len(categorized_jobs['completed_paid'])
        unsuccessful_count = len(categorized_jobs['unsuccessful'])
        total_jobs = successful_count + unsuccessful_count
        
        success_rate = (successful_count / total_jobs * 100) if total_jobs > 0 else 0
        avg_sale = (total_amount / successful_count) if successful_count > 0 else 0
        
        # Check doc checks
        all_doc_checks_pass = all(job['doc_checks'] for job in categorized_jobs['completed_paid'])
        
        # Calculate commission
        commission = self.calculate_commission(total_net_profit, scheme, all_doc_checks_pass)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commission Sheet"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write header section
        ws['A1'] = 'WEEKLY COMMISSION'
        ws['F1'] = 'WEEK ENDING'
        ws['G1'] = end_date
        
        ws['A2'] = 'BOOKED JOBS'
        ws['B2'] = total_jobs
        
        ws['A3'] = 'SUCCESSFUL'
        ws['B3'] = successful_count
        ws['D3'] = 'Tier 1'
        ws['E3'] = '<$25000'
        ws['F3'] = '5%'
        
        ws['A4'] = 'UNSUCCESSFUL'
        ws['B4'] = unsuccessful_count
        ws['D4'] = 'Tier 2'
        ws['E4'] = '>= $25000'
        ws['F4'] = '10%'
        
        ws['A5'] = 'SUCCESSFUL (%)'
        ws['B5'] = f'{success_rate:.1f}%'
        
        ws['A7'] = 'AVERAGE SALE'
        ws['B7'] = avg_sale
        ws['D7'] = 'NET PROFIT'
        ws['E7'] = total_net_profit
        ws['F7'] = total_net_profit
        
        ws['D8'] = 'COMMISSION - PAY OUT'
        ws['E8'] = commission
        
        # Job details section
        row = 12
        ws[f'A{row}'] = 'JOB  DETAILS'
        
        row += 1
        headers = ['JOB STATUS', 'DATE', 'JOB #', 'SUBURB', 'AMOUNT EXC. GST', 
                   'MATERIALS', 'Merchant Fees', 'NET PROFIT', 'PAID',
                   'BEFORE', 'AFTER', 'RECEIPT', 'DESCRIPTION', 'SIGNED', 
                   'EMAILED', 'DESCRIPTION', 'SIGNED', 'EMAILED', '5 Star Review',
                   '', 'EFTPOS', 'CASH', 'Payment Plan']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        # Write completed & paid jobs
        row += 1
        ws[f'A{row}'] = 'COMPLETED & PAID JOBS'
        row += 1
        
        for idx, job in enumerate(categorized_jobs['completed_paid'], start=1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = job['date']
            ws[f'C{row}'] = job['job_number']
            ws[f'D{row}'] = job['suburb']
            ws[f'E{row}'] = job['amount_exc_gst']
            ws[f'F{row}'] = job['materials']
            ws[f'G{row}'] = job['merchant_fees']
            ws[f'H{row}'] = job['net_profit']
            ws[f'I{row}'] = 'Card'
            # Photo/doc check columns
            for col in range(10, 19):
                ws.cell(row=row, column=col).value = 1 if job['doc_checks'] else 0
            row += 1
        
        # Write awaiting payment jobs
        row += 2
        ws[f'A{row}'] = 'CURRENT JOBS COMPLETED (AWAITING PAYMENT)'
        row += 1
        
        for idx, job in enumerate(categorized_jobs['awaiting_payment'], start=1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = job['date']
            ws[f'C{row}'] = job['job_number']
            ws[f'D{row}'] = job['suburb']
            ws[f'E{row}'] = job['amount_exc_gst']
            ws[f'F{row}'] = job['materials']
            ws[f'G{row}'] = job['merchant_fees']
            ws[f'H{row}'] = job['net_profit']
            ws[f'I{row}'] = 'Card'
            row += 1
        
        # Write unsuccessful jobs
        row += 2
        ws[f'A{row}'] = 'UNSUCCESSFUL JOBS'
        row += 1
        
        for idx, job in enumerate(categorized_jobs['unsuccessful'], start=1):
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = job['date']
            ws[f'C{row}'] = job['job_number']
            ws[f'D{row}'] = job['suburb']
            ws[f'E{row}'] = job['amount_exc_gst']
            ws[f'F{row}'] = job['materials']
            ws[f'G{row}'] = job['merchant_fees']
            ws[f'H{row}'] = job['net_profit']
            ws[f'I{row}'] = 'N/A'
            row += 1
        
        # Format currency columns
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.column in [5, 6, 7, 8, 21, 22]:  # Amount columns
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


def main():
    st.set_page_config(page_title="Commission Sheet Generator", layout="wide")
    
    st.title("ServiceTitan Commission Sheet Generator")
    
    st.session_state.st_client = helpers.get_client('foxtrotwhiskey')

    # Initialize session state
    if 'st_client' not in st.session_state:
        st.warning("⚠️ ServiceTitanClient not initialized. Please set st.session_state.st_client")
        return
    
    calculator = CommissionCalculator(st.session_state.st_client)
    
    # Sidebar for inputs
    with st.sidebar:
        st.header("Configuration")
        
        # Date range
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", datetime.now() - timedelta(days=7))
        with col2:
            end_date = st.date_input("End Date", datetime.now())
        
        # Commission scheme
        scheme = st.selectbox(
            "Commission Scheme",
            options=['scheme1', 'scheme2'],
            format_func=lambda x: "Scheme 1: 10% after $25k (all or nothing)" if x == 'scheme1' 
                                  else "Scheme 2: 5% before, 10% after $25k"
        )
        
        # Load technicians
        if st.button("🔄 Load Technicians", use_container_width=True):
            with st.spinner("Fetching technicians..."):
                try:
                    techs = calculator.fetch_technicians()
                    st.session_state.technicians = techs
                    st.success(f"Loaded {len(techs)} technicians")
                except Exception as e:
                    st.error(f"Error loading technicians: {str(e)}")
        
        # Technician selection
        if 'technicians' in st.session_state:
            tech_options = {f"{t['name']} (ID: {t['id']})": t for t in st.session_state.technicians}
            selected_tech_name = st.selectbox("Select Technician", options=list(tech_options.keys()))
            selected_tech = tech_options[selected_tech_name]
        else:
            st.info("Load technicians to continue")
            return
        
        # Generate report
        if st.button("Generate Report", type="primary", use_container_width=True):
            with st.spinner("Fetching jobs..."):
                try:
                    jobs = calculator.fetch_jobs_for_technician(
                        selected_tech['id'],
                        start_date.isoformat(),
                        end_date.isoformat()
                    )
                    st.session_state.jobs = jobs
                    st.session_state.categorized = calculator.process_jobs(jobs)
                    st.success(f"Loaded {len(jobs)} jobs")
                except Exception as e:
                    st.error(f"Error fetching jobs: {str(e)}")
    
    # Main content area
    if 'categorized' not in st.session_state:
        st.info("👈 Configure settings and generate a report to begin")
        return
    
    categorized = st.session_state.categorized
    
    # Calculate summary stats
    completed_paid = categorized['completed_paid']
    total_net_profit = sum(job['net_profit'] for job in completed_paid)
    total_amount = sum(job['amount_exc_gst'] for job in completed_paid)
    successful_count = len(completed_paid)
    unsuccessful_count = len(categorized['unsuccessful'])
    total_jobs = successful_count + unsuccessful_count
    
    success_rate = (successful_count / total_jobs * 100) if total_jobs > 0 else 0
    avg_sale = (total_amount / successful_count) if successful_count > 0 else 0
    
    all_doc_checks_pass = all(job['doc_checks'] for job in completed_paid)
    commission = calculator.calculate_commission(total_net_profit, scheme, all_doc_checks_pass)
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Net Profit", f"${total_net_profit:,.2f}")
    with col2:
        st.metric("Commission", f"${commission:,.2f}")
    with col3:
        st.metric("Success Rate", f"{success_rate:.1f}%")
    with col4:
        st.metric("Average Sale", f"${avg_sale:,.2f}")
    
    # Doc checks warning
    if not all_doc_checks_pass:
        st.warning("⚠️ Not all doc checks passed. Commission may be affected.")
    
    # Tabs for different job categories
    tab1, tab2, tab3 = st.tabs(["✅ Completed & Paid", "⏳ Awaiting Payment", "❌ Unsuccessful"])
    
    with tab1:
        st.subheader(f"Completed & Paid Jobs ({len(completed_paid)})")
        if completed_paid:
            df = pd.DataFrame(completed_paid)
            st.dataframe(
                df[['date', 'job_number', 'suburb', 'amount_exc_gst', 'materials', 'net_profit', 'doc_checks']],
                use_container_width=True
            )
            
            # Materials input section
            st.subheader("Manual Materials Input")
            st.info("Enter materials cost for each job. Leave blank to use existing values.")
            
            materials_input = {}
            for job in completed_paid:
                col1, col2, col3 = st.columns([2, 2, 1])
                with col1:
                    st.text(f"Job #{job['job_number']}")
                with col2:
                    material_cost = st.number_input(
                        f"Materials",
                        min_value=0.0,
                        value=float(job['materials']),
                        step=10.0,
                        key=f"mat_{job['job_number']}",
                        label_visibility="collapsed"
                    )
                    materials_input[job['job_number']] = material_cost
                with col3:
                    net = job['amount_exc_gst'] - material_cost
                    st.text(f"Net: ${net:,.2f}")
    
    with tab2:
        awaiting = categorized['awaiting_payment']
        st.subheader(f"Awaiting Payment ({len(awaiting)})")
        if awaiting:
            df = pd.DataFrame(awaiting)
            st.dataframe(
                df[['date', 'job_number', 'suburb', 'amount_exc_gst', 'net_profit']],
                use_container_width=True
            )
    
    with tab3:
        unsuccessful = categorized['unsuccessful']
        st.subheader(f"Unsuccessful Jobs ({len(unsuccessful)})")
        if unsuccessful:
            df = pd.DataFrame(unsuccessful)
            st.dataframe(
                df[['date', 'job_number', 'suburb', 'amount_exc_gst']],
                use_container_width=True
            )
    
    # Export button
    st.divider()
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("📥 Export to Excel", type="primary", use_container_width=True):
            excel_file = calculator.generate_excel(
                selected_tech['name'],
                categorized,
                start_date.isoformat(),
                end_date.isoformat(),
                scheme,
                materials_input if 'materials_input' in locals() else None
            )
            
            filename = f"Commission_{selected_tech['name']}_{start_date}_{end_date}.xlsx"
            st.download_button(
                label="⬇️ Download Excel",
                data=excel_file,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )


if __name__ == "__main__":
    main()